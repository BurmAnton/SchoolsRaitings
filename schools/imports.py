import math

from django.db import IntegrityError
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook
from django.core.exceptions import ObjectDoesNotExist

from email_validator import validate_email, EmailNotValidError

from .models import School, SchoolCloster, TerAdmin, SchoolType
from django.db.models import Q


SCHOOL_LEVELS = {
    '1 — 4 классы': 'S',
    '1 — 9 классы': 'M',
    '5 — 11 классы': 'MG',
    '10 — 11 классы': 'G',
    '1 — 11 классы': 'A',
}


def get_sheet(form):
    workbook = load_workbook(form.cleaned_data['import_file'])
    sheet = workbook.active
    return sheet


def cheak_col_match(sheet, fields_names_set):
    i = 0
    col_count = sheet.max_column
    sheet_fields = []
    sheet_col = {}
    if sheet[f"A2"].value is None:
        return {"status": "Error", "error_type": "EmptySheet"}
    try:
        for col_header in range(1, col_count+1):
            if sheet.cell(row=1,column=col_header).value is not None:
                sheet_fields.append(sheet.cell(row=1,column=col_header).value)
                sheet_col[col_header] = sheet.cell(row=1,column=col_header).value
        missing_fields = []
        for field in fields_names_set:
            if field not in sheet_fields:
                missing_fields.append(field)
        if len(missing_fields) != 0:
            return {"status": "Error", "error_type": "MissingFieldsError", "missing_fields": missing_fields}
    except IndexError:
        return {"status": "Error", "error_type": IndexError}
    return {"status": "OK", "sheet_col": sheet_col}


def load_worksheet_dict(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            snils = sheet[f"A{row}"].value
            if snils != None:
                cell_value = sheet.cell(row=row,column=col).value
                try: cell_value = str(math.floor(cell_value))
                except (ValueError, TypeError): pass
                sheet_dict[fields_names_set[col]].append(cell_value)
    return sheet_dict


def schools(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return ['Import', 'IndexError']
 
    #Требуемые поля таблицы
    fields_names = {
        "ID в АИС \"Кадры в образовании\"", 
        "Полное наименование",
        "Сокращенное наименование",
        "Уровень образования",
        "Номер школы",
        "Тип школы",
        "Email",
        "ТУ/ДО",
        "Населённый пункт"
    }
    
    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names["status"] == "Error":
        return cheak_col_names

    sheet = load_worksheet_dict(sheet, cheak_col_names["sheet_col"])

    missing_fields = []
    add_schools = []
    update_schools = []
    schools_id = School.objects.all().values_list("ais_id", flat=True)
    for row in range(len(sheet['ID в АИС \"Кадры в образовании\"'])):
        responce = load_school(sheet, row, schools_id)
        if responce[0] == 'MissingField':
            missing_fields.append(responce)
        elif responce[1]:
            update_schools.append(responce[2])
        else: 
            add_schools.append(responce[2])
    School.objects.bulk_create(add_schools, batch_size=50)
    School.objects.bulk_update(
        update_schools, 
        [
            "ter_admin", "email", "name", 
            "short_name", "city", "number", 
            "school_type", "closter", "ed_level"
        ], 
        50
    )

    return {
        "status": "OK",
        "added_schools": len(add_schools),
        "updated_schools": len(update_schools),
        "missing_fields": missing_fields
    }


def is_missing(field):
    if field not in ["", None]:
        return field.strip()
    return True

    
def load_school(sheet, row, id_list):
    missing_fields = []

    ais_id = is_missing(sheet["ID в АИС \"Кадры в образовании\""][row])
    if ais_id == True or not(ais_id.isnumeric()): 
        missing_fields.append("ID в АИС \"Кадры в образовании\"")
    else: ais_id = int(ais_id)
    email = is_missing(sheet["Email"][row])
    if email == True:
        missing_fields.append("Email")
    name = is_missing(sheet["Полное наименование"][row])
    if name == True: missing_fields.append("Полное наименование")
    short_name = is_missing(sheet["Сокращенное наименование"][row])
    if short_name == True: missing_fields.append("Сокращенное наименование")
    city = is_missing(sheet["Населённый пункт"][row])
    if city == True: missing_fields.append("Населённый пункт")
    number = is_missing(sheet["Номер школы"][row])
    if number == True: number = None
    ed_level = is_missing(sheet["Уровень образования"][row])
    if ed_level == True or ed_level not in SCHOOL_LEVELS.keys(): 
        #missing_fields.append("Уровень образования")
        pass
    else:
        ed_level = SCHOOL_LEVELS[ed_level]
    
    ter_admin = is_missing(sheet["ТУ/ДО"][row])
    try: ter_admin = TerAdmin.objects.get(name=ter_admin)
    except ObjectDoesNotExist: missing_fields.append("ТУ/ДО")
    school_type = is_missing(sheet["Тип школы"][row])
    try: school_type = SchoolType.objects.get(Q(short_name=school_type) | Q(name=school_type))
    except ObjectDoesNotExist: school_type = None
    closter = is_missing(sheet["Кластер"][row])
    try: closter = SchoolCloster.objects.get(name=closter)
    except ObjectDoesNotExist: closter = None

    if len(missing_fields) > 0:
        return ['MissingField', missing_fields, row+2]
    

    is_exist = ais_id in id_list
    if is_exist: school = School.objects.get(ais_id=ais_id)
    else: school = School(ais_id=ais_id)
    school.ter_admin = ter_admin
    school.email = email
    school.name = name
    school.short_name = short_name
    school.city = city
    school.number = number
    school.school_type = school_type
    school.closter = closter
    school.ed_level = ed_level

    return [
        'OK', is_exist, school
    ]