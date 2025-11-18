from django.contrib import admin
from django.urls import reverse
from django_admin_listfilter_dropdown.filters import (
    RelatedDropdownFilter, ChoiceDropdownFilter, DropdownFilter,
)
from django.utils.safestring import mark_safe
from django.contrib.admin import SimpleListFilter

from users.models import User
from reports.admin_utils import ColumnWidthMixin, add_custom_admin_css
from reports.models import Year

from .models import TerAdmin, SchoolType, School, SchoolCloster, QuestionCategory, Question


# Кастомный фильтр для архивных школ
class ArchivedFilter(SimpleListFilter):
    title = 'Архивные школы'
    parameter_name = 'is_archived'
    
    def lookups(self, request, model_admin):
        return (
            ('all', 'Все школы'),
            ('active', 'Только активные'),
            ('archived', 'Только архивные'),
        )
    
    def queryset(self, request, queryset):
        if self.value() == 'archived':
            return queryset.filter(is_archived=True)
        elif self.value() == 'active' or self.value() is None:
            return queryset.filter(is_archived=False)
        elif self.value() == 'all':
            return queryset
        return queryset
    
    def choices(self, changelist):
        # Определяем, является ли текущий выбор значением по умолчанию (None)
        is_default = self.value() is None
        
        for lookup, title in self.lookup_choices:
            # Специальная логика для опции "Только активные" - она выбрана по умолчанию
            if lookup == 'active' and is_default:
                yield {
                    'selected': True,
                    'query_string': changelist.get_query_string(remove=[self.parameter_name]),
                    'display': title,
                }
            else:
                yield {
                    'selected': self.value() == str(lookup),
                    'query_string': changelist.get_query_string({self.parameter_name: lookup}),
                    'display': title,
                }


# Register your models here.
@admin.register(TerAdmin)
class TerAdminAdmin(admin.ModelAdmin):
    search_fields = ['name', ]

    list_display = ['name', 'representatives_list']
    
    def representatives_list(self, obj):
        if obj.representatives.count() == 0:
            return "-"
        return ", ".join([repr.email for repr in obj.representatives.all()])
    representatives_list.short_description = "Представители"

    autocomplete_fields = ['representatives', ]


@admin.register(SchoolCloster)
class SchoolClosterAdmin(admin.ModelAdmin):
    pass


@admin.register(SchoolType)
class SchoolTypeAdmin(admin.ModelAdmin):
    list_display = ('name', 'short_name')
    


@admin.register(School)
@add_custom_admin_css
class SchoolAdmin(ColumnWidthMixin, admin.ModelAdmin):
    list_per_page = 50
    actions = ['export_schools', 'toggle_archive_status']
    
    # Настройки ширины столбцов
    column_width_settings = {
        'ais_id': 'column-width-sm column-align-center',
        '__str__': 'column-width-xl column-truncate',
        'number': 'column-width-sm column-align-center',
        'reports_page_link': 'column-width-sm column-align-center',
        'ter_admin': 'column-width-md column-truncate',
        'closter_field': 'column-width-md column-truncate',
        'ed_level': 'column-width-sm column-align-center',
        'email': 'column-width-md column-truncate',
        'principal': 'column-width-md column-truncate',
        'is_archived': 'column-width-sm column-align-center',
    }
    
    class Media:
        js = ('admin/js/column_width.js',)

    def toggle_archive_status(self, request, queryset):
        """Переключить статус архивации для выбранных школ"""
        archived_count = 0
        unarchived_count = 0
        
        for school in queryset:
            # Переключаем статус на противоположный
            school.is_archived = not school.is_archived
            school.save()
            
            if school.is_archived:
                archived_count += 1
            else:
                unarchived_count += 1
        
        message_parts = []
        if archived_count > 0:
            message_parts.append(f"Архивировано школ: {archived_count}")
        if unarchived_count > 0:
            message_parts.append(f"Разархивировано школ: {unarchived_count}")
        
        message = ", ".join(message_parts)
        self.message_user(request, message)
        
    toggle_archive_status.short_description = "Переключить статус архивации"

    def export_schools(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, Alignment

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        headers = [
            "ID в АИС \"Кадры в образовании\"",
            "Полное наименование",
            "Сокращенное наименование",
            "Уровень образования",
            "Кластер",
            "Номер школы",
            "Тип школы",
            "Email",
            "ТУ/ДО",
            "Населённый пункт",
            "Архивная школа",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        # Write data
        ed_levels = {
            'A': "1 — 11 классы",
            'M': "1 — 9 классы",
            'S': "1 — 4 классы",
            'G': "10 — 11 классы",
            'MG': "5 — 11 классы",
            None: "",
            'True': "",
        }
        for row, school in enumerate(queryset, 2):
            
            data = [
                school.ais_id,
                school.name,
                school.short_name,
                ed_levels[school.ed_level],
                school.closter,
                school.number,
                school.school_type,
                school.email,
                school.ter_admin,
                school.city,
                "Да" if school.is_archived else "Нет",
            ]
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col)
                if value is not None and value != "'True":
                    cell.value = str(value)
                cell.alignment = Alignment(horizontal='center')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="schools.xlsx"'
        workbook.save(response)
        return response
    export_schools.short_description = "Экспортировать выбранные школы"
    
    list_display = [
        'ais_id',
        '__str__', 
        'number',
        'reports_page_link',
        'ter_admin',
        'closter_field',
        'ed_level', 
        'email',
        'principal',
        'is_archived',
    ]
    fields = [
        'ais_id',
        'name',
        'short_name',
        'email',
        'city',
        'number',
        
        'school_type',
        'closter',
        'ed_level',
        'ter_admin', 
        'principal',
        'is_archived',
    ]

    def get_list_display(self, request):
        """Динамически определяет список отображаемых полей в зависимости от прав пользователя"""
        list_display = super().get_list_display(request)
        
        # Если пользователь не суперадмин, убираем ссылку на личный кабинет
        if not request.user.is_superuser:
            list_display = [field for field in list_display if field != 'reports_page_link']
        
        return list_display

    def reports_page_link(self, obj):
        return mark_safe(f'<a href="{reverse("reports", kwargs={"school_id": obj.id})}">Личный кабинет</a>')
    reports_page_link.short_description = "Личный кабинет"

    def principal_phone(self, obj):
        if not obj.principal or obj.principal.phone_number is None:
            return "-"
        return obj.principal.phone_number
    principal_phone.short_description = "Телефон директора"

    def principal_email(self, obj):
        if not obj.principal or obj.principal.email is None:
            return "-"
        return obj.principal.email
    principal_email.short_description = "Email директора"

    def closter_field(self, obj):
        return obj.closter
    closter_field.admin_order_field = 'closter'
    closter_field.short_description = 'Кластер'

    autocomplete_fields = ['principal', 'ter_admin', ]
    search_fields = [
        "ais_id", "name", "short_name",
        "email", "city", "number",
    ]
    list_filter = [
        ('ed_level', ChoiceDropdownFilter),
        ('school_type', RelatedDropdownFilter),
        ('ter_admin', RelatedDropdownFilter),
        ('closter', RelatedDropdownFilter),
        ArchivedFilter,
    ]
    readonly_fields = ['ais_id']
    def get_readonly_fields(self, request, obj=None):
        readonly_fields = super().get_readonly_fields(request, obj)
        
        # If user is TerAdmin representative
        if not request.user.is_superuser and hasattr(request.user, 'ter_admin') and request.user.ter_admin.exists():
            # ТУ/ДО может редактировать только кластер и уровень образования
            # Все остальные поля всегда readonly для ТУ/ДО
            editable_fields = ['closter', 'ed_level']
            
            # Создаем список всех полей модели кроме editable_fields
            readonly_fields_for_teradmin = [field.name for field in self.model._meta.fields 
                                         if field.name not in editable_fields]
                
            # Check current year status
            try:
                current_year = Year.objects.get(is_current=True)
                if current_year.status != 'updating':
                    # Если статус года не 'updating', то все поля readonly
                    return [field.name for field in self.model._meta.fields]
                else:
                    # Если статус года 'updating', то всё кроме кластера и уровня образования readonly
                    return readonly_fields_for_teradmin
            except Year.DoesNotExist:
                # If no current year, default to not editable
                return [field.name for field in self.model._meta.fields]
                
        return readonly_fields
    
    def has_change_permission(self, request, obj=None):
        # Superusers can always change
        if request.user.is_superuser:
            return True
            
        # TerAdmin representatives can only change if current year status is 'updating'
        if hasattr(request.user, 'ter_admin') and request.user.ter_admin.exists():
            try:
                current_year = Year.objects.get(is_current=True)
                return current_year.status == 'updating'
            except Year.DoesNotExist:
                return False
        
        return super().has_change_permission(request, obj)

    def has_view_permission(self, request, obj=None):
        # Суперадмины могут видеть все школы
        if request.user.is_superuser:
            return True
            
        # Представители ТУ/ДО могут видеть только свои школы
        if request.user.groups.filter(name='Представитель ТУ/ДО').exists():
            if obj is None:  # При просмотре списка школ
                return True
            # При просмотре конкретной школы проверяем, принадлежит ли она ТУ/ДО пользователя
            return obj.ter_admin in request.user.ter_admin.all()
            
        return super().has_view_permission(request, obj)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        
        # Если пользователь суперадмин, он может видеть все школы
        if request.user.is_superuser:
            return qs
            
        # Для представителей ТУ/ДО показываем только их школы
        if request.user.groups.filter(name='Представитель ТУ/ДО').exists():
            qs = qs.filter(ter_admin__in=request.user.ter_admin.all())
            
        # Применяем фильтр по архивным школам
        is_archived_param = request.GET.get('is_archived')
        if is_archived_param is None or is_archived_param == 'active':
            return qs.filter(is_archived=False)
        elif is_archived_param == 'archived':
            return qs.filter(is_archived=True)
        return qs

    def save_model(self, request, obj, form, change):
        if obj.principal and obj.email:
            obj.principal.email = obj.email
            obj.principal.save()
        super().save_model(request, obj, form, change)

@admin.register(QuestionCategory)
class QuestionCategoryAdmin(admin.ModelAdmin):
    pass

@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ['category', 'is_resolved', 'short_question', 'created_at']
    list_filter = [
        ('category', RelatedDropdownFilter), 
        'is_resolved',
        'is_visible',
    ]