from decimal import Decimal
from django.db.models import Sum, Avg, Max, Count
from common.utils import get_cache_key
from django.http import HttpResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

from reports.models import Answer, Report, SchoolReport, Section, SectionSreport, Field, Option, Year
from reports.utils import count_points, count_points_field
from schools.models import TerAdmin, School, SchoolCloster

def calculate_stats_and_section_data(f_years, reports, sections, s_reports):
    stats = {}
    section_data = {}
    
    # Get distinct sections by number
    distinct_sections = sections.distinct('number').order_by('number')
    
    for year in f_years:
        stats[year] = {
            "green_zone": SchoolReport.objects.filter(zone="G", report__year=year).count(),
            "yellow_zone": SchoolReport.objects.filter(zone="Y", report__year=year).count(),
            "red_zone": SchoolReport.objects.filter(zone="R", report__year=year).count(),
        }
        section_data[year] = []
        
        for section in distinct_sections:
            # Get all sections with the same number for this year
            year_sections = Section.objects.filter(
                number=section.number,
                report__in=reports,
                report__year=year
            )
            
            # Calculate points for all fields in these sections
            points_sum = Answer.objects.filter(
                question__sections__in=year_sections,
                s_report__report__year=year
            ).aggregate(points_sum=Sum("points"))["points_sum"] or Decimal(0)
            
            section_data[year].append(points_sum)
            
            # Initialize stats for this section
            stats[year][section.name] = {
                "green_zone": 0,
                "yellow_zone": 0,
                "red_zone": 0,
            }
            
            # Calculate zone stats
            s_reports_year = s_reports.filter(report__year=year)
            for s_report in s_reports_year:
                color = count_section_points(s_report, section)
                if color == "G":
                    stats[year][section.name]["green_zone"] += 1
                elif color == "Y":
                    stats[year][section.name]["yellow_zone"] += 1
                elif color == "R":
                    stats[year][section.name]["red_zone"] += 1
                    
    return stats, section_data

def calculate_stats(year, s_reports_year, sections):
    stats = {}
    # Initialize stats
    overall_stats = {
        "green_zone": [0, "0.0%"],
        "yellow_zone": [0, "0.0%"], 
        "red_zone": [0, "0.0%"]
    }

    s_reports_count = s_reports_year.count()

    # Create lookup dict for section stats
    for section in sections:
        stats[section.number] = {
            "green_zone": [0, "0.0%"],
            "yellow_zone": [0, "0.0%"],
            "red_zone": [0, "0.0%"]
        }

    # Calculate all stats in a single pass through the data
    for s_report in s_reports_year:
        # Overall stats
        if s_report.zone in ["G", "Y", "R"]:
            zone = {"G": "green_zone", "Y": "yellow_zone", "R": "red_zone"}[s_report.zone]
            overall_stats[zone][0] += 1

        # Section stats
        for section_sreport in s_report.sections.all():
            try:
                if section_sreport.zone in ["G", "Y", "R"]:
                    zone = {"G": "green_zone", "Y": "yellow_zone", "R": "red_zone"}[section_sreport.zone]
                    stats[section_sreport.section.number][zone][0] += 1
            except:
                continue

    # Calculate all percentages
    if s_reports_count > 0:
        # Overall percentages
        for zone in ["green_zone", "yellow_zone", "red_zone"]:
            overall_stats[zone][1] = f'{overall_stats[zone][0] / s_reports_count * 100:.1f}%'
            
        # Section percentages
        for section_stats in stats.values():
            for zone in ["green_zone", "yellow_zone", "red_zone"]:
                section_stats[zone][1] = f'{section_stats[zone][0] / s_reports_count * 100:.1f}%'

    return stats, overall_stats

def generate_ter_admins_report_csv(year, schools, s_reports):
    import xlsxwriter
    from django.http import HttpResponse
    from django.db.models import Sum, Prefetch
    from reports.models import Section, Field, Answer
    from io import BytesIO

    # Create workbook
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)

    # Define formats
    formats = {
        'header': workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'text_wrap': True, 'border': 1
        }),
        'cell': workbook.add_format({
            'align': 'center', 'valign': 'vcenter',
            'text_wrap': True, 'border': 1
        }),
        'red': workbook.add_format({
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
            'border': 1, 'bg_color': '#FFC7CE'
        }),
        'yellow': workbook.add_format({
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
            'border': 1, 'bg_color': '#FFEB9C'
        }),
        'green': workbook.add_format({
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
            'border': 1, 'bg_color': '#C6EFCE'
        })
    }

    # Create main worksheet
    worksheet = workbook.add_worksheet("Сводный отчет")

    # Get sections and prefetch related data
    sections = Section.objects.filter(report__year=year).distinct('number').order_by('number')
    
    # Prefetch answers for all sections
    sections_with_answers = sections.prefetch_related(
        Prefetch(
            'fields',
            queryset=Field.objects.prefetch_related(
                Prefetch(
                    'answers',
                    queryset=Answer.objects.filter(s_report__in=s_reports),
                    to_attr='prefetched_answers'
                )
            )
        )
    )

    # Write headers
    header = ['ТУ/ДО', 'Уровень образования', 'Школа', 'Итого баллов']
    for section in sections:
        header.append(f"{section.number}. {section.name}")

    # Add zone headers
    zone_headers = [
        ('Кол-во/доля критериев школы в зелёной зоне', ['Количество', 'Доля']),
        ('Кол-во/доля критериев школы в жёлтой зоне', ['Количество', 'Доля']), 
        ('Кол-во/доля критериев школы в красной зоне', ['Количество', 'Доля'])
    ]

    # Write header row and set column widths
    for col, value in enumerate(header):
        worksheet.write(0, col, value, formats['header'])
        width = min(max(len(value) * 1.1, 20), 50)
        worksheet.set_column(col, col, width)

    col = len(header)
    for title, columns in zone_headers:
        worksheet.merge_range(0, col, 0, col + 1, title, formats['header'])
        header.extend(columns)
        col += 2

    # Set row height
    max_text_len = max(len(str(cell)) for cell in header)
    num_lines = (max_text_len / 15) + 1
    worksheet.set_row(0, num_lines * 15)

    # Write data rows
    row_num = 1
    for school in schools:
        school_reports = s_reports.filter(school=school, report__year=year)
        if not school_reports.exists():
            continue

        for report in school_reports:
            # Basic school info
            row = [
                str(school.ter_admin),
                school.get_ed_level_display(),
                school.__str__(),  # Школа будет окрашена в цвет зоны
                report.points
            ]

            # Add section points
            for section in sections_with_answers:
                points = sum(
                    answer.points 
                    for field in section.fields.all() 
                    for answer in field.prefetched_answers 
                    if answer.s_report_id == report.id
                )
                row.append(points)

            # Add zone stats
            for zone in ['G', 'Y', 'R']:
                count = 0
                total = 0
                for section in sections_with_answers:
                    for field in section.fields.all():
                        for answer in field.prefetched_answers:
                            if answer.s_report_id == report.id:
                                total += 1
                                if answer.zone == zone:
                                    count += 1
                row.extend([count, f"{count/total*100:.1f}%" if total > 0 else "0.0%"])

            # Write row with appropriate zone formatting
            for col, value in enumerate(row):
                if col == 2:  # School column
                    format_key = {'G': 'green', 'Y': 'yellow', 'R': 'red'}.get(report.zone, 'cell')
                    worksheet.write(row_num, col, value, formats[format_key])
                else:
                    worksheet.write(row_num, col, value, formats['cell'])
            row_num += 1

    # Create worksheets for each section
    for section in sections_with_answers:
        section_worksheet = workbook.add_worksheet(f"Раздел {section.number}")
        
        # Write section headers
        section_header = ['ТУ/ДО', 'Уровень образования', 'Школа', 'Баллы раздела']
        fields = sorted(section.fields.all(), key=lambda x: [int(n) for n in str(x.number).split('.')])
        
        for field in fields:
            section_header.append(f"{field.number}. {field.name}")

        # Write header row and set column widths
        for col, value in enumerate(section_header):
            section_worksheet.write(0, col, value, formats['header'])
            width = min(max(len(value) * 1.1, 20), 50)
            section_worksheet.set_column(col, col, width)

        # Write data rows
        row_num = 1
        for school in schools:
            school_reports = s_reports.filter(school=school, report__year=year)
            if not school_reports.exists():
                continue

            for report in school_reports:
                # Basic school info
                row = [
                    str(school.ter_admin),
                    school.get_ed_level_display(),
                    school.__str__()  # Школа будет окрашена в цвет зоны
                ]

                # Get section points
                section_points = sum(
                    answer.points 
                    for field in section.fields.all() 
                    for answer in field.prefetched_answers 
                    if answer.s_report_id == report.id
                )
                row.append(section_points)

                # Add field values
                for field in fields:
                    answer = next(
                        (a for a in field.prefetched_answers if a.s_report_id == report.id),
                        None
                    )
                    row.append(answer.points if answer else 0)

                # Write row with appropriate zone formatting
                for col, value in enumerate(row):
                    if col == 2:  # School column
                        format_key = {'G': 'green', 'Y': 'yellow', 'R': 'red'}.get(report.zone, 'cell')
                        section_worksheet.write(row_num, col, value, formats[format_key])
                    else:
                        section_worksheet.write(row_num, col, value, formats['cell'])
                row_num += 1

    # Close workbook
    workbook.close()
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ter_admins_report_{year.year}.xlsx'
    
    return response

def write_section_details(worksheet, row_num, sections, s_reports, formats):
    for section in sections:
        # Write section header
        worksheet.write(row_num, 0, f'Раздел {section.number}. {section.name}', formats['header'])
        row_num += 1

        # Write field headers
        headers = ['ТУ/ДО', 'Уровень образования', 'Школа', 'Всего баллов']
        sections_objs = Section.objects.filter(name=section.name)
        fields = Field.objects.filter(sections__in=sections_objs).distinct('number').prefetch_related('answers')
        fields = sorted(fields, key=lambda x: [int(n) for n in str(x.number).split('.')])
        for field in fields:
            headers.append(f'{field.number}. {field.name}')

        for col, header in enumerate(headers):
            worksheet.write(row_num, col, header, formats['header'])
        row_num += 1

        # Write school data for each field
        for s_report in s_reports:
            school = s_report.school
            row = [
                str(school.ter_admin),
                school.get_ed_level_display(),
                school.__str__(),
                s_report.points
            ]
            
            for field in fields:
                answer = Answer.objects.filter(question=field, s_report=s_report).first()
                points = answer.points if answer else "-"
                zone = answer.zone if answer else 'W'
                format_key = {'R': 'red', 'Y': 'yellow', 'G': 'green'}.get(zone, 'cell')
                row.append(points)
                worksheet.write(row_num, len(row)-1, points, formats[format_key])
            
            for col in range(4):
                worksheet.write(row_num, col, row[col], formats['cell'])
            row_num += 1
        
        row_num += 1

    return row_num

def write_summary_rows(worksheet, row_num, s_reports, sections, formats):
    # Write totals row
    total_row = ['Итого', '', '', 0]
    total_sum = 0
    
    for section in sections:
        sections_objs = Section.objects.filter(name=section.name)
        fields = Field.objects.filter(sections__in=sections_objs).distinct('number').prefetch_related('answers')
        fields = sorted(fields, key=lambda x: [int(n) for n in str(x.number).split('.')])
        section_total = Answer.objects.filter(
            question__in=fields, 
            s_report__in=s_reports
        ).aggregate(Sum('points'))['points__sum'] or 0
        total_row.append(section_total)
        total_sum += section_total
    
    total_row[3] = total_sum
    
    for col, value in enumerate(total_row):
        worksheet.write(row_num, col, value, formats['header'])
    
    return row_num + 1


def count_section_points(s_report, section):
    from reports.models import Answer
    try:
        section_obj = Section.objects.filter(number=section.number, report=s_report.report).first()
        points__sum = Answer.objects.filter(question__in=section_obj.fields.all(), s_report=s_report).aggregate(Sum('points'))['points__sum']
        if points__sum < section_obj.yellow_zone_min:
            return "R"
        elif points__sum >= section_obj.green_zone_min:
            return "G"
        return "Y"
    except:
        return "R"

def generate_school_report_csv(year, school, s_reports, sections):
    import xlsxwriter
    from django.http import HttpResponse
    from io import BytesIO

    # Create workbook
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    
    # Define formats
    formats = {
        'header': workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'text_wrap': True, 'border': 1
        }),
        'cell': workbook.add_format({
            'align': 'center', 'valign': 'vcenter',
            'text_wrap': True, 'border': 1
        }),
        'red': workbook.add_format({
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
            'border': 1, 'bg_color': '#FFC7CE'
        }),
        'yellow': workbook.add_format({
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
            'border': 1, 'bg_color': '#FFEB9C'
        }),
        'green': workbook.add_format({
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
            'border': 1, 'bg_color': '#C6EFCE'
        }),
        'section_header': workbook.add_format({
            'bold': True, 'align': 'left', 'valign': 'vcenter',
            'text_wrap': True, 'border': 1, 'bg_color': '#E7E6E6'
        })
    }

    # Create main worksheet
    worksheet = workbook.add_worksheet("Сводный отчет")

    # Write headers
    header = ['Год', 'ТУ/ДО', 'Уровень образования', 'Школа', 'Итого баллов']
    
    # Add section headers
    for section in sections:
        header.append(f"{section.number}. {section.name}")

    # Write header row and set column widths
    for col, value in enumerate(header):
        worksheet.write(0, col, value, formats['header'])
        width = min(max(len(value) * 1.1, 20), 50)
        worksheet.set_column(col, col, width)

    # Write summary data rows
    row_num = 1
    for s_report in s_reports:
        # Basic report info
        row = [
            str(s_report.report.year),
            str(school.ter_admin),
            school.get_ed_level_display(),
            school.__str__(),
            s_report.points
        ]

        # Add section points
        for section in sections:
            section_report = SectionSreport.objects.filter(
                s_report=s_report,
                section__number=section.number
            ).first()
            points = section_report.points if section_report else 0
            row.append(points)

        # Write row data with appropriate formatting
        for col, value in enumerate(row):
            format_key = 'cell'
            if col == 4:  # Total points column
                format_key = {'R': 'red', 'Y': 'yellow', 'G': 'green'}.get(s_report.zone, 'cell')
            elif col >= 5:  # Section columns
                section = sections[col-5]
                section_report = SectionSreport.objects.filter(
                    s_report=s_report,
                    section__number=section.number
                ).first()
                if section_report:
                    format_key = {'R': 'red', 'Y': 'yellow', 'G': 'green'}.get(section_report.zone, 'cell')
            
            worksheet.write(row_num, col, value, formats[format_key])

        row_num += 1

    # Create detailed worksheets for each section
    for section in sections:
        worksheet = workbook.add_worksheet(f"Раздел {section.number}")
        
        # Write section header
        worksheet.merge_range(0, 0, 0, 4, f"{section.number}. {section.name}", formats['section_header'])
        
        # Write column headers
        headers = ['Год', 'ТУ/ДО', 'Школа', 'Показатель', 'Баллы']
        for col, header in enumerate(headers):
            worksheet.write(1, col, header, formats['header'])
            worksheet.set_column(col, col, 20)
        
        row_num = 2
        for s_report in s_reports:
            # Get all fields (questions) for this section
            fields = Field.objects.filter(sections=section).distinct('number').order_by('number')
            
            for field in fields:
                answer = Answer.objects.filter(
                    question=field,
                    s_report=s_report
                ).first()
                
                row = [
                    str(s_report.report.year),
                    str(school.ter_admin),
                    school.__str__(),
                    f"{section.number}.{field.number}. {field.name}",
                    answer.points if answer else 0
                ]
                
                # Write row with appropriate formatting
                format_key = 'cell'
                if answer:
                    format_key = {'R': 'red', 'Y': 'yellow', 'G': 'green'}.get(answer.zone, 'cell')
                
                for col, value in enumerate(row):
                    worksheet.write(row_num, col, value, formats['cell' if col < 4 else format_key])
                
                row_num += 1
            
            # Add a blank row between years
            row_num += 1

    # Close workbook
    workbook.close()
    output.seek(0)
    
    # Generate response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="school_report_{school.id}_{year}.xlsx"'
    
    return response

def generate_closters_report_csv(year, schools, s_reports):
    import xlsxwriter
    from django.http import HttpResponse
    from io import BytesIO

    # Create workbook
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    
    # Define formats
    formats = {
        'header': workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'text_wrap': True, 'border': 1
        }),
        'cell': workbook.add_format({
            'align': 'center', 'valign': 'vcenter',
            'text_wrap': True, 'border': 1
        }),
        'red': workbook.add_format({
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
            'border': 1, 'bg_color': '#FFC7CE'
        }),
        'yellow': workbook.add_format({
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
            'border': 1, 'bg_color': '#FFEB9C'
        }),
        'green': workbook.add_format({
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
            'border': 1, 'bg_color': '#C6EFCE'
        }),
        'section_header': workbook.add_format({
            'bold': True, 'align': 'left', 'valign': 'vcenter',
            'text_wrap': True, 'border': 1, 'bg_color': '#E7E6E6'
        })
    }

    # Create summary worksheet
    worksheet = workbook.add_worksheet("Итоговые баллы")

    # Write school info headers for summary
    headers = ['Показатель', 'Максимум', 'Среднее']
    for s_report in s_reports:
        school = s_report.school
        headers.append(f"{school.ter_admin} - {school}")

    # Write headers and set column widths
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, formats['header'])
        width = min(max(len(header) * 1.2, 15), 40)
        worksheet.set_column(col, col, width)

    # Get sections
    sections = Section.objects.filter(report__year=year).distinct('number').order_by('number')
    
    # Write data for each section in summary
    row_num = 1
    for section in sections:
        # Write section name
        worksheet.write(row_num, 0, f"{section.number}. {section.name}", formats['section_header'])
        
        # Calculate max and average for section
        section_stats = SectionSreport.objects.filter(
            s_report__in=s_reports,
            section__number=section.number
        ).aggregate(
            max_points=Max('points'),
            avg_points=Avg('points')
        )
        
        # Write max and average
        worksheet.write(row_num, 1, section_stats['max_points'] or 0, formats['cell'])
        worksheet.write(row_num, 2, round(section_stats['avg_points'] or 0, 1), formats['cell'])
        
        # Write points for each school
        for col, s_report in enumerate(s_reports, 3):
            section_report = SectionSreport.objects.filter(
                s_report=s_report,
                section__number=section.number
            ).first()
            points = section_report.points if section_report else 0
            format_key = {'R': 'red', 'Y': 'yellow', 'G': 'green'}.get(
                section_report.zone if section_report else '', 'cell'
            )
            
            worksheet.write(row_num, col, points, formats[format_key])
        
        row_num += 1

    # Create worksheet for each section
    for section in sections:
        worksheet = workbook.add_worksheet(f"Раздел {section.number}")
        
        # Write headers for section worksheet
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, formats['header'])
            worksheet.set_column(col, col, width)

        row_num = 1
        
        # Write section total row
        worksheet.write(row_num, 0, f"{section.number}. {section.name}", formats['section_header'])
        
        # Calculate and write section totals
        section_stats = SectionSreport.objects.filter(
            s_report__in=s_reports,
            section__number=section.number
        ).aggregate(
            max_points=Max('points'),
            avg_points=Avg('points')
        )
        
        worksheet.write(row_num, 1, section_stats['max_points'] or 0, formats['cell'])
        worksheet.write(row_num, 2, round(section_stats['avg_points'] or 0, 1), formats['cell'])
        
        for col, s_report in enumerate(s_reports, 3):
            section_report = SectionSreport.objects.filter(
                s_report=s_report,
                section__number=section.number
            ).first()
            
            points = section_report.points if section_report else 0
            format_key = {'R': 'red', 'Y': 'yellow', 'G': 'green'}.get(
                section_report.zone if section_report else '', 'cell'
            )
            
            worksheet.write(row_num, col, points, formats[format_key])
        
        row_num += 2  # Add blank row after section total

        # Write field data
        fields = Field.objects.filter(sections=section).distinct('number').order_by('number')
        for field in fields:
            worksheet.write(row_num, 0, f"{section.number}.{field.number}. {field.name}", formats['cell'])
            
            # Calculate max and average for field
            field_stats = Answer.objects.filter(
                question=field,
                s_report__in=s_reports
            ).aggregate(
                max_points=Max('points'),
                avg_points=Avg('points')
            )
            
            # Write max and average
            worksheet.write(row_num, 1, field_stats['max_points'] or 0, formats['cell'])
            worksheet.write(row_num, 2, round(field_stats['avg_points'] or 0, 1), formats['cell'])
            
            # Write points for each school
            for col, s_report in enumerate(s_reports, 3):
                answer = Answer.objects.filter(
                    question=field,
                    s_report=s_report
                ).first()
                
                points = answer.points if answer else 0
                format_key = {'R': 'red', 'Y': 'yellow', 'G': 'green'}.get(
                    answer.zone if answer else '', 'cell'
                )
                
                worksheet.write(row_num, col, points, formats[format_key])
            
            row_num += 1

    # Close workbook
    workbook.close()
    output.seek(0)
    
    # Generate response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="closters_report_{year}.xlsx"'
    
    return response

def calculate_answers_distribution(year, schools_reports, sections, ter_admin=None, closter=None, ed_level=None):
    """
    Рассчитывает распределение вариантов ответов для каждого показателя
    с разбивкой по ТУ/ДО, кластерам и уровням образования
    
    :param year: год отчета
    :param schools_reports: QuerySet с отчетами школ
    :param sections: разделы отчета
    :param ter_admin: фильтр по территориальному управлению
    :param closter: фильтр по кластеру
    :param ed_level: фильтр по уровню образования
    :return: словарь со статистикой по распределению ответов
    """
    # Словарь, в котором будем хранить результаты
    result = {
        'overall_stats': {},           # Общая статистика по всем школам
        'ter_admin_stats': {},         # Статистика по ТУ/ДО
        'closter_stats': {},           # Статистика по кластерам
        'ed_level_stats': {},          # Статистика по уровням образования
        'field_school_counts': {},     # Количество школ, заполнивших показатель
        'last_monitoring_date': str(year)   # Дата последнего мониторинга
    }
    
    # Подготавливаем списки для группировки и фильтрации
    ter_admin_ids = set()
    closter_ids = set()
    ed_levels = set()
    
    # Получаем отчеты школ с фильтрацией
    filtered_schools_reports = []
    
    for report in schools_reports:
        school = report.school
        if not school:
            continue
        
        # Применяем фильтры
        if ter_admin and school.ter_admin_id != ter_admin.id:
            continue
        
        if closter and school.closter_id != closter.id:
            continue
        
        if ed_level and school.ed_level != ed_level:
            continue
        
        # Добавляем данные в списки для группировки
        ter_admin_ids.add(school.ter_admin_id)
        if school.closter_id:
            closter_ids.add(school.closter_id)
        ed_levels.add(school.ed_level)
        
        # Добавляем отчет в список
        filtered_schools_reports.append(report)

    # Обрабатываем каждую секцию
    for section in sections:
        section_id = str(section.id)
        
        # Инициализируем секцию в словарях статистики
        if section_id not in result['overall_stats']:
            result['overall_stats'][section_id] = {}
        
        # Обрабатываем каждое поле секции
        for field in section.fields.all():
            field_id = str(field.id)
            
            # Инициализируем поле в словаре общей статистики
            if field_id not in result['overall_stats'][section_id]:
                result['overall_stats'][section_id][field_id] = {}
            
            # Словари для подсчета по разным разрезам
            overall_counts = {}  # Общий счетчик по вариантам ответов
            ter_admin_counts = {ter_id: {} for ter_id in ter_admin_ids}  # Счетчики по ТУ/ДО
            closter_counts = {closter_id: {} for closter_id in closter_ids}  # Счетчики по кластерам
            ed_level_counts = {level: {} for level in ed_levels}  # Счетчики по уровням образования
            
            # Количество школ, заполнивших данный показатель
            field_schools_count = 0
            
            # Обрабатываем каждую школу
            for report in filtered_schools_reports:
                school = report.school
                if not school:
                    continue
                
                # Получаем ответ на вопрос
                answer = None
                for a in report.answers.all():
                    if a.question_id == field.id:
                        answer = a
                        break
                
                if not answer:
                    continue
                
                # Увеличиваем счетчик школ
                field_schools_count += 1
                
                # Получаем идентификатор и имя варианта ответа
                option_id = None
                option_name = None
                zone = answer.zone if answer.zone else 'N'
                
                # Определяем вариант ответа в зависимости от типа поля
                if field.answer_type == 'LST':  # Выбор из списка
                    if answer.option:
                        option_id = str(answer.option.id)
                        option_name = answer.option.name
                elif field.answer_type == 'BL':  # Бинарный ответ (да/нет)
                    option_id = 'yes' if answer.bool_value else 'no'
                    option_name = 'Да' if answer.bool_value else 'Нет'
                elif field.answer_type == 'MULT':  # Множественный выбор
                    # Для множественного выбора создаем отдельные счетчики для каждого выбранного варианта
                    for opt in answer.selected_options.all():
                        option_id = str(opt.id)
                        option_name = opt.name
                        
                        # Обновляем общую статистику
                        if option_id not in overall_counts:
                            overall_counts[option_id] = {
                                'count': 0,
                                'percentage': 0,
                                'option_name': option_name,
                                'zone': zone
                            }
                        overall_counts[option_id]['count'] += 1
                        
                        # Обновляем статистику по ТУ/ДО
                        if school.ter_admin_id:
                            if option_id not in ter_admin_counts[school.ter_admin_id]:
                                ter_admin_counts[school.ter_admin_id][option_id] = {
                                    'count': 0,
                                    'percentage': 0,
                                    'option_name': option_name,
                                    'zone': zone
                                }
                            ter_admin_counts[school.ter_admin_id][option_id]['count'] += 1
                        
                        # Обновляем статистику по кластерам
                        if school.closter_id:
                            if option_id not in closter_counts[school.closter_id]:
                                closter_counts[school.closter_id][option_id] = {
                                    'count': 0,
                                    'percentage': 0,
                                    'option_name': option_name,
                                    'zone': zone
                                }
                            closter_counts[school.closter_id][option_id]['count'] += 1
                        
                        # Обновляем статистику по уровням образования
                        if option_id not in ed_level_counts[school.ed_level]:
                            ed_level_counts[school.ed_level][option_id] = {
                                'count': 0,
                                'percentage': 0,
                                'option_name': option_name,
                                'zone': zone
                            }
                        ed_level_counts[school.ed_level][option_id]['count'] += 1
                    
                    # Пропускаем остальную логику для множественного выбора, т.к. уже обработали все варианты
                    continue
                elif field.answer_type in ['NMBR', 'PRC']:  # Числовой ответ или процент
                    # Для числовых полей разбиваем на диапазоны
                    option_id = f'range_{get_range_for_number(answer.number_value, field)}'
                    option_name = f'{answer.number_value}'
                
                # Если не удалось определить вариант ответа, пропускаем
                if option_id is None:
                    continue
            
            # Сохраняем количество школ, заполнивших показатель
            result['field_school_counts'][field_id] = field_schools_count
            
            # Если нет данных по показателю, переходим к следующему
            if field_schools_count == 0:
                continue
            
            # Рассчитываем проценты для общей статистики
            for option_id, data in overall_counts.items():
                data['percentage'] = round(data['count'] * 100 / field_schools_count, 1)
                result['overall_stats'][section_id][field_id][option_id] = data
            
            # Рассчитываем проценты для статистики по ТУ/ДО
            for ter_id, options in ter_admin_counts.items():
                # Подсчитываем количество школ в данном ТУ/ДО, заполнивших показатель
                ter_schools_count = sum(1 for report in filtered_schools_reports 
                                     if report.school and report.school.ter_admin_id == ter_id)
                
                if ter_schools_count == 0:
                    continue
                
                # Инициализируем структуру данных, если нужно
                ter_admin = TerAdmin.objects.get(id=ter_id)
                ter_admin_name = ter_admin.name if ter_admin else str(ter_id)
                
                if ter_id not in result['ter_admin_stats']:
                    result['ter_admin_stats'][ter_id] = {
                        'name': ter_admin_name,
                    }
                
                if section_id not in result['ter_admin_stats'][ter_id]:
                    result['ter_admin_stats'][ter_id][section_id] = {}
                
                if field_id not in result['ter_admin_stats'][ter_id][section_id]:
                    result['ter_admin_stats'][ter_id][section_id][field_id] = {}
                
                # Рассчитываем проценты для каждого варианта ответа
                for option_id, data in options.items():
                    data['percentage'] = round(data['count'] * 100 / ter_schools_count, 1)
                    result['ter_admin_stats'][ter_id][section_id][field_id][option_id] = data
            
            # Рассчитываем проценты для статистики по кластерам
            for closter_id, options in closter_counts.items():
                # Подсчитываем количество школ в данном кластере, заполнивших показатель
                closter_schools_count = sum(1 for report in filtered_schools_reports 
                                     if report.school and report.school.closter_id == closter_id)
                
                if closter_schools_count == 0:
                    continue
                
                # Инициализируем структуру данных, если нужно
                closter = SchoolCloster.objects.get(id=closter_id)
                closter_name = closter.name if closter else str(closter_id)
                
                if closter_id not in result['closter_stats']:
                    result['closter_stats'][closter_id] = {
                        'name': closter_name,
                    }
                
                if section_id not in result['closter_stats'][closter_id]:
                    result['closter_stats'][closter_id][section_id] = {}
                
                if field_id not in result['closter_stats'][closter_id][section_id]:
                    result['closter_stats'][closter_id][section_id][field_id] = {}
                
                # Рассчитываем проценты для каждого варианта ответа
                for option_id, data in options.items():
                    data['percentage'] = round(data['count'] * 100 / closter_schools_count, 1)
                    result['closter_stats'][closter_id][section_id][field_id][option_id] = data
            
            # Рассчитываем проценты для статистики по уровням образования
            for ed_level, options in ed_level_counts.items():
                # Подсчитываем количество школ данного уровня, заполнивших показатель
                ed_level_schools_count = sum(1 for report in filtered_schools_reports 
                                     if report.school and report.school.ed_level == ed_level)
                
                if ed_level_schools_count == 0:
                    continue
                
                # Инициализируем структуру данных, если нужно
                ed_level_name = dict(School.SCHOOL_LEVELS).get(ed_level, str(ed_level))
                
                if ed_level not in result['ed_level_stats']:
                    result['ed_level_stats'][ed_level] = {
                        'name': ed_level_name,
                    }
                
                if section_id not in result['ed_level_stats'][ed_level]:
                    result['ed_level_stats'][ed_level][section_id] = {}
                
                if field_id not in result['ed_level_stats'][ed_level][section_id]:
                    result['ed_level_stats'][ed_level][section_id][field_id] = {}
                
                # Рассчитываем проценты для каждого варианта ответа
                for option_id, data in options.items():
                    data['percentage'] = round(data['count'] * 100 / ed_level_schools_count, 1)
                    result['ed_level_stats'][ed_level][section_id][field_id][option_id] = data
    
    return result

def generate_answers_distribution_excel(year, schools_reports, sections, stats):
    """
    Создает Excel-файл с распределением вариантов ответов по показателям
    
    :param year: год отчета
    :param schools_reports: отчеты школ
    :param sections: разделы отчета
    :param stats: статистические данные
    :return: HttpResponse с Excel-файлом
    """
    # Создаем workbook
    workbook = Workbook()
    
    # Настройка форматирования для разных зон
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    # Добавляем общий лист с информацией
    info_sheet = workbook.active
    info_sheet.title = "Общая информация"
    
    # Заголовок
    info_sheet['A1'] = f"Распределение вариантов ответов по показателям за {year} год"
    info_sheet['A1'].font = Font(size=14, bold=True)
    info_sheet.merge_cells('A1:F1')
    
    # Информация о фильтрах
    info_sheet['A3'] = "Дата формирования:"
    info_sheet['B3'] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    
    info_sheet['A4'] = "Всего образовательных организаций:"
    info_sheet['B4'] = len(schools_reports)
    
    row = 6
    
    # Добавляем информацию о зонах
    info_sheet['A6'] = "Цветовые обозначения:"
    info_sheet['A7'] = "Красная зона"
    info_sheet['A7'].fill = red_fill
    
    info_sheet['A8'] = "Желтая зона"
    info_sheet['A8'].fill = yellow_fill
    
    info_sheet['A9'] = "Зеленая зона"
    info_sheet['A9'].fill = green_fill
    
    # Обрабатываем каждый раздел
    for section in sections:
        # Создаем лист для раздела
        sheet_name = f"{section.number}. {section.name[:20]}"
        sheet = workbook.create_sheet(sheet_name)
        
        # Заголовок раздела
        sheet['A1'] = f"{section.number}. {section.name}"
        sheet['A1'].font = Font(size=14, bold=True)
        sheet.merge_cells('A1:F1')
        
        row = 3
        
        # Добавляем данные по каждому полю раздела
        for field in section.fields.all():
            # Заголовок поля
            header = f"{section.number}.{field.number}. {field.name}"
            sheet[f'A{row}'] = header
            sheet[f'A{row}'].font = Font(size=12, bold=True)
            sheet.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Количество школ, заполнивших это поле
            sheet[f'A{row}'] = f"Всего школ: {stats['field_school_counts'].get(field.id, 0)}"
            sheet.merge_cells(f'A{row}:F{row}')
            row += 2
            
            # Заголовки таблицы для общей статистики
            sheet[f'A{row}'] = "Общее распределение вариантов ответов"
            sheet[f'A{row}'].font = Font(bold=True)
            sheet.merge_cells(f'A{row}:F{row}')
            row += 1
            
            sheet[f'A{row}'] = "Вариант ответа"
            sheet[f'B{row}'] = "Количество"
            sheet[f'C{row}'] = "Процент"
            sheet[f'D{row}'] = "Зона"
            
            for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                sheet[cell].font = Font(bold=True)
            row += 1
            
            # Данные общей статистики
            overall_stats = stats['overall_stats'].get(str(section.id), {}).get(str(field.id), {})
            for option_id, option_data in overall_stats.items():
                sheet[f'A{row}'] = option_data['option_name']
                sheet[f'B{row}'] = option_data['count']
                sheet[f'C{row}'] = f"{option_data['percentage']}%"
                sheet[f'D{row}'] = option_data['zone']
                
                # Применяем цветовую маркировку
                zone_cell = sheet[f'D{row}']
                if option_data['zone'] == 'R':
                    for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                        sheet[cell].fill = red_fill
                elif option_data['zone'] == 'Y':
                    for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                        sheet[cell].fill = yellow_fill
                elif option_data['zone'] == 'G':
                    for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                        sheet[cell].fill = green_fill
                
                row += 1
            
            row += 2
            
            # Создаем таблицы для ТУ/ДО, кластеров и уровней образования
            # Таблица по ТУ/ДО
            if stats['ter_admin_stats']:
                sheet[f'A{row}'] = "Распределение по ТУ/ДО"
                sheet[f'A{row}'].font = Font(bold=True)
                sheet.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Собираем все варианты ответов для этого поля
                options = set()
                ter_admins = []
                
                for ter_admin_id, ter_admin_data in stats['ter_admin_stats'].items():
                    if ter_admin_data.get(str(section.id), {}).get(str(field.id)):
                        ter_admins.append(ter_admin_id)
                        for option_id in ter_admin_data[str(section.id)][str(field.id)].keys():
                            options.add(option_id)
                
                # Заголовки таблицы
                sheet[f'A{row}'] = "ТУ/ДО"
                col = 1
                options_list = sorted(list(options))
                option_cols = {}  # Для сопоставления вариантов ответов с столбцами
                
                for option_id in options_list:
                    option_name = ""
                    # Найдем имя варианта ответа из любой статистики, где он есть
                    for ter_admin_id in ter_admins:
                        ter_admin_data = stats['ter_admin_stats'].get(ter_admin_id, {})
                        if (ter_admin_data.get(str(section.id), {}).get(str(field.id), {}).get(option_id)):
                            option_name = ter_admin_data[str(section.id)][str(field.id)][option_id]['option_name']
                            break
                    
                    col_letter = get_column_letter(col + 1)
                    sheet[f'{col_letter}{row}'] = option_name
                    sheet[f'{col_letter}{row}'].font = Font(bold=True)
                    option_cols[option_id] = col_letter
                    col += 1
                
                row += 1
                
                # Данные по ТУ/ДО
                for ter_admin_id in sorted(ter_admins):
                    sheet[f'A{row}'] = ter_admin_id
                    
                    for option_id in options_list:
                        if (stats['ter_admin_stats'].get(ter_admin_id, {}).get(str(section.id), {}).get(str(field.id), {}).get(option_id)):
                            option_data = stats['ter_admin_stats'][ter_admin_id][str(section.id)][str(field.id)][option_id]
                            col_letter = option_cols[option_id]
                            sheet[f'{col_letter}{row}'] = f"{option_data['percentage']}%"
                            
                            # Применяем цветовую маркировку
                            if option_data['zone'] == 'R':
                                sheet[f'{col_letter}{row}'].fill = red_fill
                            elif option_data['zone'] == 'Y':
                                sheet[f'{col_letter}{row}'].fill = yellow_fill
                            elif option_data['zone'] == 'G':
                                sheet[f'{col_letter}{row}'].fill = green_fill
                        else:
                            col_letter = option_cols[option_id]
                            sheet[f'{col_letter}{row}'] = "0%"
                    
                    row += 1
                
                row += 2
            
            # Таблица по кластерам
            if stats['closter_stats']:
                sheet[f'A{row}'] = "Распределение по кластерам"
                sheet[f'A{row}'].font = Font(bold=True)
                sheet.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Собираем все варианты ответов для этого поля
                options = set()
                closters = []
                
                for closter_id, closter_data in stats['closter_stats'].items():
                    if closter_data.get(str(section.id), {}).get(str(field.id)):
                        closters.append(closter_id)
                        for option_id in closter_data[str(section.id)][str(field.id)].keys():
                            options.add(option_id)
                
                # Заголовки таблицы
                sheet[f'A{row}'] = "Кластер"
                col = 1
                options_list = sorted(list(options))
                option_cols = {}  # Для сопоставления вариантов ответов с столбцами
                
                for option_id in options_list:
                    option_name = ""
                    # Найдем имя варианта ответа из любой статистики, где он есть
                    for closter_id in closters:
                        closter_data = stats['closter_stats'].get(closter_id, {})
                        if (closter_data.get(str(section.id), {}).get(str(field.id), {}).get(option_id)):
                            option_name = closter_data[str(section.id)][str(field.id)][option_id]['option_name']
                            break
                    
                    col_letter = get_column_letter(col + 1)
                    sheet[f'{col_letter}{row}'] = option_name
                    sheet[f'{col_letter}{row}'].font = Font(bold=True)
                    option_cols[option_id] = col_letter
                    col += 1
                
                row += 1
                
                # Данные по кластерам
                for closter_id in sorted(closters):
                    sheet[f'A{row}'] = closter_id
                    
                    for option_id in options_list:
                        if (stats['closter_stats'].get(closter_id, {}).get(str(section.id), {}).get(str(field.id), {}).get(option_id)):
                            option_data = stats['closter_stats'][closter_id][str(section.id)][str(field.id)][option_id]
                            col_letter = option_cols[option_id]
                            sheet[f'{col_letter}{row}'] = f"{option_data['percentage']}%"
                            
                            # Применяем цветовую маркировку
                            if option_data['zone'] == 'R':
                                sheet[f'{col_letter}{row}'].fill = red_fill
                            elif option_data['zone'] == 'Y':
                                sheet[f'{col_letter}{row}'].fill = yellow_fill
                            elif option_data['zone'] == 'G':
                                sheet[f'{col_letter}{row}'].fill = green_fill
                        else:
                            col_letter = option_cols[option_id]
                            sheet[f'{col_letter}{row}'] = "0%"
                    
                    row += 1
                
                row += 2
            
            # Таблица по уровням образования
            if stats['ed_level_stats']:
                sheet[f'A{row}'] = "Распределение по уровням образования"
                sheet[f'A{row}'].font = Font(bold=True)
                sheet.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Собираем все варианты ответов для этого поля
                options = set()
                ed_levels = []
                
                for ed_level, ed_level_data in stats['ed_level_stats'].items():
                    if ed_level_data.get(str(section.id), {}).get(str(field.id)):
                        ed_levels.append(ed_level)
                        for option_id in ed_level_data[str(section.id)][str(field.id)].keys():
                            options.add(option_id)
                
                # Заголовки таблицы
                sheet[f'A{row}'] = "Уровень образования"
                col = 1
                options_list = sorted(list(options))
                option_cols = {}  # Для сопоставления вариантов ответов с столбцами
                
                for option_id in options_list:
                    option_name = ""
                    # Найдем имя варианта ответа из любой статистики, где он есть
                    for ed_level in ed_levels:
                        ed_level_data = stats['ed_level_stats'].get(ed_level, {})
                        if (ed_level_data.get(str(section.id), {}).get(str(field.id), {}).get(option_id)):
                            option_name = ed_level_data[str(section.id)][str(field.id)][option_id]['option_name']
                            break
                    
                    col_letter = get_column_letter(col + 1)
                    sheet[f'{col_letter}{row}'] = option_name
                    sheet[f'{col_letter}{row}'].font = Font(bold=True)
                    option_cols[option_id] = col_letter
                    col += 1
                
                row += 1
                
                # Данные по уровням образования
                for ed_level in sorted(ed_levels):
                    sheet[f'A{row}'] = ed_level
                    
                    for option_id in options_list:
                        if (stats['ed_level_stats'].get(ed_level, {}).get(str(section.id), {}).get(str(field.id), {}).get(option_id)):
                            option_data = stats['ed_level_stats'][ed_level][str(section.id)][str(field.id)][option_id]
                            col_letter = option_cols[option_id]
                            sheet[f'{col_letter}{row}'] = f"{option_data['percentage']}%"
                            
                            # Применяем цветовую маркировку
                            if option_data['zone'] == 'R':
                                sheet[f'{col_letter}{row}'].fill = red_fill
                            elif option_data['zone'] == 'Y':
                                sheet[f'{col_letter}{row}'].fill = yellow_fill
                            elif option_data['zone'] == 'G':
                                sheet[f'{col_letter}{row}'].fill = green_fill
                        else:
                            col_letter = option_cols[option_id]
                            sheet[f'{col_letter}{row}'] = "0%"
                    
                    row += 1
                
                row += 2
    
    # Настройка автоширины столбцов
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в ответ HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="answers_distribution_{year}.xlsx"'
    workbook.save(response)
    
    return response

def get_schools_reports(year, user):
    """
    Получает данные отчетов школ за указанный год
    
    :param year: год отчета
    :param user: текущий пользователь
    :return: словарь с данными отчетов школ
    """
    from schools.models import TerAdmin, School
    
    # Получаем объект года или используем строковое представление
    if not isinstance(year, Year):
        try:
            year = Year.objects.get(year=year)
        except Year.DoesNotExist:
            return {}
    
    # Определяем доступные территориальные управления для пользователя
    is_ter_admin = TerAdmin.objects.filter(representatives=user).exists()
    
    if is_ter_admin:
        ter_admins = TerAdmin.objects.filter(representatives=user)
    else:
        ter_admins = TerAdmin.objects.all()
    
    # Получаем школы из доступных территориальных управлений
    schools = School.objects.filter(ter_admin__in=ter_admins, is_archived=False)
    
    # Получаем отчеты за указанный год
    reports = Report.objects.filter(year=year, is_published=True)
    
    # Получаем отчеты школ
    schools_reports = SchoolReport.objects.filter(
        report__in=reports,
        school__in=schools,
        status='D'  # Только принятые отчеты
    ).select_related(
        'school',
        'report',
        'school__ter_admin',
        'school__closter',
    ).prefetch_related(
        'answers',
        'sections__section__fields',
        'sections__section',
    )
    
    # Преобразуем в словарь для удобного доступа
    result = {}
    for s_report in schools_reports:
        school_id = s_report.school.id
        if school_id not in result:
            result[school_id] = {
                'school': s_report.school,
                'sections': {},
            }
        
        # Добавляем данные по секциям
        for section in s_report.sections.all():
            section_id = str(section.section.id)
            if section_id not in result[school_id]['sections']:
                result[school_id]['sections'][section_id] = {
                    'id': section.section.id,
                    'name': section.section.name,
                    'number': section.section.number,
                    'points': section.points,
                    'zone': section.zone,
                    'fields': {},
                }
            
            # Создаем индекс для быстрого поиска ответов
            answers_lookup = {a.question_id: a for a in s_report.answers.all()}
            
            # Добавляем данные по полям
            for field in section.section.fields.all():
                field_id = str(field.id)
                answer = answers_lookup.get(field.id)
                if answer:
                    result[school_id]['sections'][section_id]['fields'][field_id] = {
                        'id': field.id,
                        'name': field.name,
                        'number': field.number,
                        'type': field.answer_type,
                        'answers': []
                    }
                    
                    # Добавляем данные об ответе в зависимости от типа поля
                    if field.answer_type == 'LST':
                        # Ответ из списка
                        result[school_id]['sections'][section_id]['fields'][field_id]['answers'].append({
                            'option_id': answer.option.id if answer.option else None,
                            'option_name': answer.option.name if answer.option else None,
                            'zone': answer.zone,
                            'points': answer.points
                        })
                    elif field.answer_type == 'BL':
                        # Бинарный ответ (да/нет)
                        result[school_id]['sections'][section_id]['fields'][field_id]['answers'].append({
                            'option_id': 'yes' if answer.bool_value else 'no',
                            'option_name': 'Да' if answer.bool_value else 'Нет',
                            'zone': answer.zone,
                            'points': answer.points
                        })
                    elif field.answer_type == 'MULT':
                        # Множественный выбор
                        for option in answer.selected_options.all():
                            result[school_id]['sections'][section_id]['fields'][field_id]['answers'].append({
                                'option_id': option.id,
                                'option_name': option.name,
                                'zone': option.zone,
                                'points': None  # В множественном выборе баллы привязаны ко всему ответу
                            })
                    elif field.answer_type in ['NMBR', 'PRC']:
                        # Числовой ответ или процент
                        result[school_id]['sections'][section_id]['fields'][field_id]['answers'].append({
                            'option_id': f'range_{get_range_for_number(answer.number_value, field)}',
                            'option_name': f'{answer.number_value}',
                            'zone': answer.zone,
                            'points': answer.points
                        })
    
    return result

def get_range_for_number(value, field):
    """Вспомогательная функция для определения диапазона числового значения"""
    # Получаем максимальное значение для данного поля
    max_value = Answer.objects.filter(question=field).aggregate(Max('number_value'))['number_value__max'] or 0
    
    if max_value == 0:
        return 1
    
    # Создаем 5 равных диапазонов
    range_size = max_value / 5
    
    # Определяем, к какому диапазону относится значение
    if value <= range_size:
        return 1
    elif value <= range_size * 2:
        return 2
    elif value <= range_size * 3:
        return 3
    elif value <= range_size * 4:
        return 4
    else:
        return 5

