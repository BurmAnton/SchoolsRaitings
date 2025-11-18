from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.core.cache import cache
from django.views.decorators.cache import cache_page
from django.db.models import Sum, Prefetch
from django.conf import settings
from django.contrib import messages
from collections import defaultdict
from django.http import HttpResponse, JsonResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from django.contrib.auth.models import Group

from dashboards import utils
from reports.models import Answer, Report, SchoolReport, Section, Field, Year, SectionSreport
from schools.models import SchoolCloster, School, TerAdmin
from common.utils import get_cache_key

# Create your views here.
@login_required
def ter_admins_reports(request):
    reports = Report.objects.all().prefetch_related('reports', 'schools', 'schools__school', 'sections')

    return render(request, "dashboards/ter_admins_reports.html", {
        'reports': reports,
    })

#@cache_page(None, key_prefix="flow")
@login_required
@csrf_exempt
def ter_admins_dash(request):
    # Check if the user is a TerAdmin representatives
    ter_admins = TerAdmin.objects.filter(representatives=request.user)

    if not ter_admins.first():
        ter_admins = TerAdmin.objects.all()

    closters = SchoolCloster.objects.all()
    ed_levels = {
        'A': "1 — 11 классы",
        'M': "1 — 9 классы",
        'S': "1 — 4 классы",
        'MG': "5 — 11 классы",
        'G': "10 — 11 классы",
    }
    years = Year.objects.all().order_by('-year')
    
    year = Year.objects.get(is_current=True)
    schools_reports = None
    sections = None
    school_reports_data = {}
    stats = {}
    overall_stats = {}
    sections_data = {}
    fields_sum_data = {}
    show_ter_status = False
    filter_params = {}
    schools = School.objects.filter(ter_admin__in=ter_admins, is_archived=False)
    reports = Report.objects.filter(year=year)
  
    if request.method == 'POST':
        year = request.POST["year"]
        year = Year.objects.get(year=year)
        show_ter_status = 'show_ter_status' in request.POST
        reports = Report.objects.filter(year=year)
        sections = Section.objects.filter(report__in=reports).distinct('number').order_by('number')
        
        # Фильтруем отчеты в зависимости от статуса согласования

        if show_ter_status:
            schools_reports = SchoolReport.objects.filter(
                report__in=reports,
                status__in=['A', 'D']  # Показываем и принятые, и на согласовании
            ).order_by('school__name').prefetch_related('answers', 'sections')
        else:
            schools_reports = SchoolReport.objects.filter(
                report__in=reports,
                status='D'  # Только принятые
            ).order_by('school__name').prefetch_related('answers', 'sections')
        
        stats, overall_stats = utils.calculate_stats(year, schools_reports, sections)
        
        # Фильтруем школы по выбранным параметрам
        ter_admins_f = request.POST.get("ter_admin", '')
        if ter_admins_f and ter_admins_f != 'all':
            schools = schools.filter(ter_admin=ter_admins_f)
            filter_params['ter_admin'] = ter_admins_f
        elif ter_admins_f == 'all':
            # Если выбрано "Все управления/департаменты", не фильтруем школы
            filter_params['ter_admin'] = 'all'
            
        closters_f = request.POST.getlist("closters")
        if closters_f:
            schools = schools.filter(closter__in=closters_f)
            filter_params['closters'] = closters_f
            
        ed_levels_f = request.POST.getlist("ed_levels")
        if ed_levels_f:
            schools = schools.filter(ed_level__in=ed_levels_f)
            filter_params['ed_levels'] = ed_levels_f
            
        # Фильтруем отчеты по отфильтрованным школам
        schools_reports = schools_reports.filter(school__in=schools)
        if 'download' in request.POST:
            return utils.generate_ter_admins_report_csv(year, schools, schools_reports)

    # Generate cache key based on filters
    cache_key = get_cache_key('ter_admins_dash', 
        year=year,
        show_ter_status=show_ter_status,
        schools=','.join(sorted(str(s.id) for s in schools)),
        reports=','.join(sorted(str(r.id) for r in reports)) if reports else '',
        ter_admin=filter_params.get('ter_admin', ''),
        closters=','.join(sorted(filter_params.get('closters', []))),
        ed_levels=','.join(sorted(filter_params.get('ed_levels', [])))
    )
    
    # Try to get cached data
    cached_data = cache.get(cache_key)
    if cached_data and reports:  # Проверяем наличие reports
        # Фильтруем отчеты по статусу согласования, если это необходимо
        if show_ter_status:
            cached_data['schools_reports'] = cached_data['schools_reports'].filter(
                status__in=['A', 'D']
            )
        else:
            cached_data['schools_reports'] = cached_data['schools_reports'].filter(
                status='D'
            )
            
        return render(request, "dashboards/ter_admins_dash.html", {
            **cached_data,
            "years": years,
            "selected_year": year,
            "filter": filter_params,
            'ter_admins': ter_admins,
            'closters': closters,
            'ed_levels': ed_levels,
            'show_ter_status': show_ter_status
        })

    # If no cached data, perform the expensive calculations
    if reports:
        schools_reports = SchoolReport.objects.filter(
            report__in=reports,
            school__in=schools,
            status__in=['A', 'D'] if show_ter_status else ['D']
        ).select_related(
            'school',
            'report',
            'school__ter_admin',
        ).prefetch_related(
            'answers',
            'sections__section__fields',
            'sections__section',
        )
    else:
        schools_reports = SchoolReport.objects.none()

    school_reports_data = {}
    fields_data = {}
    sections = Section.objects.filter(report__in=reports).distinct('number').order_by('number').prefetch_related('fields') if reports else []

    for s_report in schools_reports:
        school_reports_data[s_report.id] = {
            'green_zone_answers': 0,
            'yellow_zone_answers': 0,
            'red_zone_answers': 0,
            'answers': 0
        }
        
        # Create lookup dict for answers
        answer_lookup = {a.question_id: a for a in s_report.answers.all()}
        
        for section in s_report.sections.all():
            # Используем название раздела вместо номера
            section_name = section.section.name
            if section_name not in fields_data:
                fields_data[section_name] = {
                    'fields': {}
                }
            school_reports_data[s_report.id][section_name] = {
                'points': section.points,
                'zone': section.zone,
                'fields': {}
            }
            
            fields = sorted(
                section.section.fields.all(), 
                key=lambda x: [int(n) for n in str(x.number).split('.')]
            )
            
            for field in fields:
                if field.number not in fields_data[section_name]['fields']:
                    fields_data[section_name]['fields'][field.number] = {
                        'points': 0,
                        'green_zone': 0,
                        'yellow_zone': 0,
                        'red_zone': 0
                    }
                answer = answer_lookup.get(field.id)
                if answer:
                    school_reports_data[s_report.id][section_name]['fields'][field.number] = {
                        'points': answer.points,
                        'zone': answer.zone
                    }
    
                    fields_data[section_name]['fields'][field.number]['points'] += answer.points
                    school_reports_data[s_report.id]['answers'] += 1
                    if answer.zone == 'G':
                        school_reports_data[s_report.id]['green_zone_answers'] += 1
                        fields_data[section_name]['fields'][field.number]['green_zone'] += 1
                    elif answer.zone == 'Y':
                        school_reports_data[s_report.id]['yellow_zone_answers'] += 1
                        fields_data[section_name]['fields'][field.number]['yellow_zone'] += 1
                    elif answer.zone == 'R':
                        school_reports_data[s_report.id]['red_zone_answers'] += 1
                        fields_data[section_name]['fields'][field.number]['red_zone'] += 1
                else:
                    school_reports_data[s_report.id][section_name]['fields'][field.number] = {
                        'points': 0,
                        'zone': 'W'
                    }

    stats, overall_stats = utils.calculate_stats(year, schools_reports, sections)
    sections_data = {}

    sections_data = {}
    for section in sections:
        sections_objs = Section.objects.filter(name=section.name)
        fields = Field.objects.filter(sections__in=sections_objs).distinct('number').prefetch_related('answers')
        # Используем название раздела в качестве ключа
        sections_data[section.name] = {
            'name': section.name,
            'fields': sorted(fields, key=lambda x: [int(n) for n in str(x.number).split('.')])
        }
    fields_sum_data = {}
    # fields = Field.objects.filter(sections__in=sections).distinct('number').prefetch_related('answers')
    for key, section in sections_data.items():
        for field in section['fields']:
            fields_sum_data[field.id] = field.answers.filter(s_report__in=schools_reports).aggregate(Sum('points'))['points__sum'] or 0

    # Prepare data for caching
    cache_data = {
        'reports': reports,
        'schools_reports': schools_reports,
        'sections': sections,
        'stats': stats,
        'overall_stats': overall_stats,
        'school_reports_data': school_reports_data,
        'fields_data': fields_data,
        'sections_data': sections_data,
        'fields_sum_data': fields_sum_data
    }

    # Cache the data for 10 minutes
    cache.set(cache_key, cache_data, timeout=600)

    # Проверяем, является ли пользователь администратором или представителем МинОбр
    try:
        mo_group = Group.objects.get(name='Представитель МинОбр')
        is_mo_or_admin = request.user.is_superuser or request.user.groups.filter(id=mo_group.id).exists()
    except Group.DoesNotExist:
        is_mo_or_admin = request.user.is_superuser

    return render(request, "dashboards/ter_admins_dash.html", {
        **cache_data,
        "years": years,
        "selected_year": year,
        "filter": filter_params,
        'ter_admins': ter_admins,
        'closters': closters,
        'ed_levels': ed_levels,
        'show_ter_status': show_ter_status
    })


@login_required
@csrf_exempt
def school_report(request):
    # Определяем, является ли пользователь директором школы
    is_school_principal = hasattr(request.user, 'school') and request.user.school
    
    # Получаем доступные ТУ/ДО
    ter_admins = TerAdmin.objects.filter(representatives=request.user).prefetch_related('schools')
    if not ter_admins.first():
        ter_admins = TerAdmin.objects.all().prefetch_related('schools')
    
    years = Year.objects.all().order_by('-year')
    
    school = None
    s_reports = None
    sections = None
    section_data = {}
    stats = {}
    filter_params = {}
    reports = None
    fields_sum_data = {}

    # Если пользователь - директор школы, автоматически используем его школу
    if is_school_principal:
        school = request.user.school
        
        # Берем последний год по умолчанию или годы из запроса
        if request.method == 'POST' and 'years' in request.POST:
            f_years = request.POST.getlist("years")
            f_years = [Year.objects.get(year=year) for year in f_years]
        else:
            current_year = years.first()
            if current_year:
                f_years = [current_year]
            else:
                f_years = []
        
        if f_years:
            reports = Report.objects.filter(year__in=f_years)
            sections = Section.objects.filter(report__in=reports).distinct('number').order_by('number')
            s_reports = SchoolReport.objects.filter(
                report__in=reports, 
                school=school, 
                status='D'
            ).order_by('report__year').prefetch_related('answers', 'sections')
            
            filter_params = {
                'years': f_years,
                'school': str(school.id),
                'ter_admin': str(school.ter_admin.id) if school.ter_admin else None
            }
            
            # Handle export request
            if request.method == 'POST' and 'download' in request.POST:
                return utils.generate_school_report_csv(f_years[-1], school, s_reports, sections)
            
            stats, section_data = utils.calculate_stats_and_section_data(f_years, reports, sections, s_reports)

            # Calculate fields sum data only for existing answers
            sections = Section.objects.filter(report__in=reports).distinct('number').order_by('number')
            fields = Field.objects.filter(sections__in=sections).distinct('number').prefetch_related('answers')
            
            for field in fields:
                answers_sum = field.answers.filter(
                    s_report__in=s_reports,
                    points__isnull=False
                ).aggregate(Sum('points'))['points__sum']
                fields_sum_data[field.id] = answers_sum if answers_sum is not None else 0
    # Если это не директор школы, обрабатываем GET и POST запросы стандартно
    else:
        # Проверяем, передан ли идентификатор школы в GET-параметрах
        if request.method == 'GET' and 'school' in request.GET:
            try:
                school_id = int(request.GET.get('school'))
                school = School.objects.get(id=school_id)
            except (ValueError, School.DoesNotExist):
                school = None

            # Если школа определена, автоматически формируем отчет
            if school:
                # Берем последний год по умолчанию
                current_year = years.first()
                if current_year:
                    f_years = [current_year]
                    reports = Report.objects.filter(year__in=f_years)
                    sections = Section.objects.filter(report__in=reports).distinct('number').order_by('number')
                    s_reports = SchoolReport.objects.filter(
                        report__in=reports, 
                        school=school, 
                        status='D'
                    ).order_by('report__year').prefetch_related('answers', 'sections')
                    
                    filter_params = {
                        'years': f_years,
                        'school': str(school.id),
                        'ter_admin': str(school.ter_admin.id) if school.ter_admin else None
                    }
                    
                    stats, section_data = utils.calculate_stats_and_section_data(f_years, reports, sections, s_reports)

                    # Calculate fields sum data only for existing answers
                    sections = Section.objects.filter(report__in=reports).distinct('number').order_by('number')
                    fields = Field.objects.filter(sections__in=sections).distinct('number').prefetch_related('answers')
                    
                    for field in fields:
                        answers_sum = field.answers.filter(
                            s_report__in=s_reports,
                            points__isnull=False
                        ).aggregate(Sum('points'))['points__sum']
                        fields_sum_data[field.id] = answers_sum if answers_sum is not None else 0

        # Обработка POST запроса (выбор фильтров)
        if request.method == 'POST' and not is_school_principal:
            school = School.objects.get(id=request.POST["school"])
            f_years = request.POST.getlist("years")
            f_years = [Year.objects.get(year=year) for year in f_years]
            reports = Report.objects.filter(year__in=f_years)
            sections = Section.objects.filter(report__in=reports).distinct('number').order_by('number')
            s_reports = SchoolReport.objects.filter(
                report__in=reports, 
                school=school, 
                status='D'
            ).order_by('report__year').prefetch_related('answers', 'sections')
            
            filter_params = {
                'years': f_years,
                'school': str(school.id),
                'ter_admin': str(school.ter_admin.id) if school.ter_admin else None
            }
            
            # Handle export request
            if 'download' in request.POST:
                return utils.generate_school_report_csv(f_years[-1], school, s_reports, sections)
                
            stats, section_data = utils.calculate_stats_and_section_data(f_years, reports, sections, s_reports)

            # Calculate fields sum data only for existing answers
            sections = Section.objects.filter(report__in=reports).distinct('number').order_by('number')
            fields = Field.objects.filter(sections__in=sections).distinct('number').prefetch_related('answers')
            
            for field in fields:
                answers_sum = field.answers.filter(
                    s_report__in=s_reports,
                    points__isnull=False
                ).aggregate(Sum('points'))['points__sum']
                fields_sum_data[field.id] = answers_sum if answers_sum is not None else 0

    # Подготавливаем список школ для выбора в зависимости от роли пользователя
    schools_for_select = []
    if is_school_principal:
        schools_for_select = [request.user.school]
    else:
        # Для остальных пользователей показываем все доступные школы
        for ter_admin in ter_admins:
            schools_for_select.extend(ter_admin.schools.filter(is_archived=False))

    return render(request, "dashboards/school_report.html", {
        "filter": filter_params,
        "ter_admins": ter_admins,
        "years": years,
        "schools": schools_for_select,
        "sections": sections,
        "section_data": section_data,
        "stats": stats,
        "fields_sum_data": fields_sum_data,
        "s_reports": s_reports,
        "dash_school": school
    })


# Диагностическая функция для отладки
@login_required
def debug_closters_report(request):
    """Диагностическая функция для выявления проблем с данными"""
    debug_info = []
    
    try:
        # === РАСШИРЕННАЯ ДИАГНОСТИКА АВТОРИЗАЦИИ ===
        debug_info.append("=== ДИАГНОСТИКА АВТОРИЗАЦИИ ===")
        debug_info.append(f"request.user: {request.user}")
        debug_info.append(f"request.user type: {type(request.user)}")
        debug_info.append(f"request.user.is_authenticated: {getattr(request.user, 'is_authenticated', 'N/A')}")
        debug_info.append(f"request.user.is_anonymous: {getattr(request.user, 'is_anonymous', 'N/A')}")
        debug_info.append(f"request.user.id: {getattr(request.user, 'id', 'N/A')}")
        debug_info.append(f"request.user.username: {getattr(request.user, 'username', 'N/A')}")
        debug_info.append(f"request.user.email: {getattr(request.user, 'email', 'N/A')}")
        debug_info.append(f"request.user.is_superuser: {getattr(request.user, 'is_superuser', 'N/A')}")
        debug_info.append(f"request.user.is_staff: {getattr(request.user, 'is_staff', 'N/A')}")
        debug_info.append(f"request.user.is_active: {getattr(request.user, 'is_active', 'N/A')}")
        
        # Проверка сессии
        debug_info.append(f"Session key: {request.session.session_key}")
        debug_info.append(f"Session data keys: {list(request.session.keys())}")
        debug_info.append(f"USER_ID in session: {request.session.get('_auth_user_id', 'NOT FOUND')}")
        
        # Проверка middleware
        debug_info.append(f"Remote IP: {request.META.get('REMOTE_ADDR', 'N/A')}")
        debug_info.append(f"HTTP_X_FORWARDED_FOR: {request.META.get('HTTP_X_FORWARDED_FOR', 'N/A')}")
        
        debug_info.append("=== ДИАГНОСТИКА БАЗЫ ДАННЫХ ===")
        
        # Проверяем количество лет
        years_count = Year.objects.count()
        debug_info.append(f"Количество лет в базе: {years_count}")
        
        if years_count > 0:
            current_year = Year.objects.filter(is_current=True).first()
            if current_year:
                debug_info.append(f"Текущий год: {current_year.year}")
            else:
                debug_info.append("Текущий год не установлен")
                
        # Проверяем ТУ/ДО
        ter_admins_count = TerAdmin.objects.count()
        debug_info.append(f"Количество ТУ/ДО: {ter_admins_count}")
        
        # Проверяем школы
        schools_count = School.objects.count()
        active_schools_count = School.objects.filter(is_archived=False).count()
        debug_info.append(f"Всего школ: {schools_count}, активных: {active_schools_count}")
        
        # Проверяем кластеры
        closters_count = SchoolCloster.objects.count()
        debug_info.append(f"Количество кластеров: {closters_count}")
        
        # Проверяем отчеты
        reports_count = Report.objects.count()
        debug_info.append(f"Количество шаблонов отчетов: {reports_count}")
        
        school_reports_count = SchoolReport.objects.count()
        debug_info.append(f"Количество отчетов школ: {school_reports_count}")
        
        # Проверяем группы пользователей
        try:
            from django.contrib.auth.models import Group
            mo_group = Group.objects.filter(name='Представитель МинОбр').first()
            if mo_group:
                debug_info.append(f"Группа 'Представитель МинОбр' существует, участников: {mo_group.user_set.count()}")
            else:
                debug_info.append("Группа 'Представитель МинОбр' не найдена")
        except Exception as e:
            debug_info.append(f"Ошибка при проверке групп: {e}")
            
        # Проверяем пользователя более детально (ИСПРАВЛЕНО для кастомной модели)
        user_identifier = getattr(request.user, 'email', None) or getattr(request.user, 'username', None)
        if request.user and request.user.is_authenticated and user_identifier:
            debug_info.append(f"✅ Пользователь корректно авторизован: {user_identifier}")
            debug_info.append(f"Суперпользователь: {request.user.is_superuser}")
            user_ter_admins = TerAdmin.objects.filter(representatives=request.user).count()
            debug_info.append(f"Пользователь представляет ТУ/ДО: {user_ter_admins}")
            
            # Проверяем группы пользователя
            user_groups = request.user.groups.all()
            debug_info.append(f"Группы пользователя: {[g.name for g in user_groups]}")
            
            # Дополнительная информация о модели пользователя
            debug_info.append(f"USERNAME_FIELD модели: {getattr(request.user.__class__, 'USERNAME_FIELD', 'username')}")
            debug_info.append(f"Email пользователя: {getattr(request.user, 'email', 'N/A')}")
            debug_info.append(f"Имя: {getattr(request.user, 'first_name', 'N/A')}")
            debug_info.append(f"Фамилия: {getattr(request.user, 'last_name', 'N/A')}")
        else:
            debug_info.append("❌ ПРОБЛЕМА: Пользователь не авторизован или отсутствует идентификатор")
            
        # === ДИАГНОСТИКА DJANGO НАСТРОЕК ===
        debug_info.append("=== ДИАГНОСТИКА DJANGO НАСТРОЕК ===")
        from django.conf import settings
        debug_info.append(f"AUTH_USER_MODEL: {getattr(settings, 'AUTH_USER_MODEL', 'НЕ УСТАНОВЛЕНО')}")
        debug_info.append(f"LOGIN_URL: {getattr(settings, 'LOGIN_URL', 'НЕ УСТАНОВЛЕНО')}")
        debug_info.append(f"DEBUG: {getattr(settings, 'DEBUG', 'НЕ УСТАНОВЛЕНО')}")
        
        # Проверяем модель пользователя
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            debug_info.append(f"Модель пользователя: {User}")
            debug_info.append(f"USERNAME_FIELD: {getattr(User, 'USERNAME_FIELD', 'username')}")
            total_users = User.objects.count()
            active_users = User.objects.filter(is_active=True).count()
            superusers = User.objects.filter(is_superuser=True).count()
            debug_info.append(f"Всего пользователей: {total_users}, активных: {active_users}, админов: {superusers}")
        except Exception as e:
            debug_info.append(f"Ошибка при проверке модели пользователя: {e}")
        
    except Exception as e:
        debug_info.append(f"Ошибка при диагностике: {str(e)}")
        import traceback
        debug_info.append(f"Traceback: {traceback.format_exc()}")
        
    return render(request, "dashboards/debug.html", {"debug_info": debug_info})


# @cache_page(60 * 5, key_prefix="closters_report")
@login_required
@csrf_exempt
def closters_report(request, year=2024):
    # ИСПРАВЛЕННАЯ ПРОВЕРКА: Для кастомной модели пользователя с email вместо username
    if not request.user or not request.user.is_authenticated:
        messages.error(request, "Ошибка авторизации: пользователь не авторизован. Пожалуйста, войдите в систему.")
        from django.shortcuts import redirect
        return redirect('login_view')
    
    # Проверяем наличие email (для кастомной модели) или username
    user_identifier = getattr(request.user, 'email', None) or getattr(request.user, 'username', None)
    if not user_identifier:
        messages.error(request, "Ошибка: отсутствует идентификатор пользователя. Обратитесь к администратору.")
        from django.shortcuts import redirect
        return redirect('login_view')
        
    filter_params = {}
    years = Year.objects.all().order_by('-year')
    if years:
        current_year = years[0]
    else:
        # Если нет годов в базе, создаем сообщение об ошибке
        messages.error(request, "В системе отсутствуют данные о годах отчетности.")
        return render(request, "dashboards/closters_report.html", {})

    # --- Инициализация переменных и справочников ---
    show_ter_status = False  # По умолчанию не показываем согласование

    # Доступные территориальные управления для пользователя
    ter_admins = TerAdmin.objects.filter(representatives=request.user)
    if not ter_admins.first():
        ter_admins = TerAdmin.objects.all()

    # Для админа и представителей МинОбр доступ ко всем ТУ/ДО
    ter_admins_for_schools = ter_admins
    try:
        # Безопасная проверка группы
        mo_group_exists = request.user.groups.filter(name='Представитель МинОбр').exists()
        iro_group = request.user.groups.filter(name='Представитель ИРО').exists()
        if request.user.is_superuser or mo_group_exists or iro_group:
            ter_admins_for_schools = TerAdmin.objects.all()
            filter_params['ter_admin'] = 'all'
    except Exception as e:
        # Логируем ошибку, но продолжаем работу
        print(f"Ошибка при проверке групп пользователя: {e}")

    # --- Проверяем наличие необходимых данных ---
    closters = SchoolCloster.objects.all()
    if not closters.exists():
        messages.warning(request, "В системе отсутствуют данные о кластерах школ.")
    
    ed_levels = {
        'A': "1 — 11 классы",
        'M': "1 — 9 классы",
        'S': "1 — 4 классы",
        'MG': "5 — 11 классы",
        'G': "10 — 11 классы",
    }

    # Базовый queryset школ (далее будет фильтроваться)
    schools = School.objects.filter(
        ter_admin__in=ter_admins_for_schools,
        is_archived=False
    ).select_related('ter_admin', 'closter')

    # --- Отладочная информация ---
    try:
        is_mo_or_admin = request.user.is_superuser or request.user.groups.filter(name='Представитель МинОбр').exists() or request.user.groups.filter(name='Представитель ИРО').exists()
    except Exception:
        is_mo_or_admin = request.user.is_superuser

    # ------------------------------------------------------------
    # Формирование данных отчёта
    # ------------------------------------------------------------

    # Определяем выбранный год (selected_year) для шаблона
    selected_year = current_year
    if request.method == 'POST':
        year_val = request.POST.get('year')
        if year_val:
            selected_year = Year.objects.filter(year=year_val).first() or current_year

    # Проверяем наличие отчетов для выбранного года
    try:
        reports_for_year = Report.objects.filter(year=selected_year)
        if not reports_for_year.exists():
            messages.warning(request, f"Отсутствуют отчеты за {selected_year.year} год.")
    except Exception as e:
        print(f"Ошибка при проверке отчетов: {e}")
        reports_for_year = Report.objects.none()

    # Если пришёл POST — применяем фильтры
    if request.method == 'POST':
        # Показывать отчёты на согласовании ТУ/ДО
        show_ter_status = 'show_ter_status' in request.POST
        filter_params['show_ter_status'] = show_ter_status

        # Фильтр по ТУ/ДО
        ter_admins_f = request.POST.get('ter_admin', '')
        if ter_admins_f:
            if ter_admins_f != 'all':
                schools = schools.filter(ter_admin_id=ter_admins_f)
                filter_params['ter_admin'] = ter_admins_f
            else:
                filter_params['ter_admin'] = 'all'

        # Фильтр по кластерам
        closters_f = request.POST.getlist('closters')
        if closters_f:
            schools = schools.filter(closter_id__in=closters_f)
            filter_params['closters'] = closters_f

        # Фильтр по уровням образования
        ed_levels_f = request.POST.getlist('ed_levels')
        if ed_levels_f:
            schools = schools.filter(ed_level__in=ed_levels_f)
            filter_params['ed_levels'] = ed_levels_f

    # Запрашиваем отчёты школ с учётом статуса согласования и отфильтрованных школ
    try:
        s_reports_qs = SchoolReport.objects.filter(
            report__year=selected_year,
            status__in=['A', 'D'] if show_ter_status else ['D'],
            school__in=schools
        ).select_related(
            'school', 'school__ter_admin', 'school__closter'
        ).prefetch_related(
            'sections__section',
            # Критично: ответы вместе с вопросом, чтобы избежать N+1 по названию поля
            Prefetch('answers', queryset=Answer.objects.select_related('question'))
        )
    except Exception as e:
        print(f"Ошибка при получении отчетов школ: {e}")
        s_reports_qs = SchoolReport.objects.none()

    # Обработка запроса на скачивание Excel
    if request.method == 'POST' and 'download' in request.POST:
        return utils.generate_closters_report_csv(selected_year, schools, s_reports_qs)

    # --- Секции и поля ---
    try:
        sections_src = Section.objects.filter(report__year=selected_year).order_by('number').prefetch_related('fields')
    except Exception as e:
        print(f"Ошибка при получении секций: {e}")
        sections_src = Section.objects.none()

    # Уникальные номера секций (на случай дублей)
    seen_numbers = set()
    sections = []  # формат [(num, name, fields_qs)]
    for sec in sections_src:
        if sec.number in seen_numbers:
            continue
        seen_numbers.add(sec.number)
        sections.append((sec.number, sec.name, sec.fields.all().order_by('number')))

    # Для графика – списки номеров и названий
    section_numbers = [str(item[0]) for item in sections]
    section_names = [f"{item[0]}. {item[1]}" for item in sections]

    # --- Формируем словарь значений по школам ---
    school_values: dict[str, list[float]] = {}
    # Подготовим шаблон списка с нулями нужной длины
    zero_template = [0.0] * len(sections)

    # Предзагрузим SectionSreport для всех отчётов, чтобы не делать N+1
    try:
        ssr_qs = SectionSreport.objects.filter(s_report__in=s_reports_qs).select_related('section')
        # Индексация: {(report_id, section_number): points}
        ssr_dict = { (ssr.s_report_id, ssr.section.number): ssr.points for ssr in ssr_qs }
    except Exception as e:
        print(f"Ошибка при получении данных секций отчетов: {e}")
        ssr_dict = {}

    for s_report in s_reports_qs:
        try:
            school_name = str(s_report.school)
            values = zero_template.copy()
            for idx, (num, _name, _fields) in enumerate(sections):
                points = ssr_dict.get((s_report.id, num), 0)
                values[idx] = float(points) if points is not None else 0.0
            school_values[school_name] = values
        except Exception as e:
            print(f"Ошибка при обработке отчета школы {s_report.school}: {e}")
            continue

    # ------------------------------------------------------------
    # Формируем контекст и рендерим шаблон
    # ------------------------------------------------------------

    context = {
        'years': years,
        'ter_admins': ter_admins,
        'closters': closters,
        'ed_levels': ed_levels,
        's_reports': s_reports_qs,
        'sections': sections,
        'schools': schools,
        'filter': filter_params,
        'show_ter_status': show_ter_status,
        'selected_year': selected_year,
        'section_names': section_names,
        'section_numbers': section_numbers,
        'school_values': school_values,
        'is_mo_or_admin': is_mo_or_admin,
    }

    return render(request, "dashboards/closters_report.html", context)


# ===== Временные заглушки для отсутствующих представлений =====
def answers_distribution_report(request):
    """
    Представление для отчета о распределении зон (красная, жёлтая, зелёная) 
    по территориальным управлениям (ТУ/ДО) с возможностью фильтрации по годам.
    """
    # Получаем доступные годы и территориальные управления для пользователя
    years = Year.objects.all().order_by('-year')
    ter_admins = TerAdmin.objects.filter(representatives=request.user)
    
    # Флаг, показывающий, что пользователь - представитель ТУ/ДО
    is_ter_admin_rep = ter_admins.exists()
    
    if not is_ter_admin_rep:
        ter_admins = TerAdmin.objects.all()
    
    # Подготовка контекста для фильтра
    filter_context = {
        'years': years,
        'ter_admins': ter_admins,
        'is_ter_admin_rep': is_ter_admin_rep
    }
    
    # Статистика по умолчанию (пустая)
    stats = []
    selected_years = []
    show_year_column = False
    show_ter_status = False  # Добавляем новый параметр для фильтра
    
    # Инициализируем переменные для статистики, используемые в Excel-выгрузке
    section_stats = []
    cluster_stats = []
    indicator_stats = []
    
    # --- DEBUG START ---
    print(f"Request method: {request.method}")
    print(f"Session selected_year_ids before processing: {request.session.get('selected_year_ids')}")
    if request.method == 'POST':
        print(f"POST data years: {request.POST.getlist('years')}")
    # --- DEBUG END ---

    # Обработка запроса с фильтрами или установка всех годов по умолчанию
    if request.method == 'POST':
        # Получаем значение фильтра "Показывать школы на согласовании ТУ/ДО"
        show_ter_status = 'show_ter_status' in request.POST
        # Сохраняем в сессию
        request.session['show_ter_status'] = show_ter_status
        
        # Получаем выбранные годы из POST-запроса
        year_ids = request.POST.getlist('years')
        if year_ids:
            selected_years = Year.objects.filter(id__in=year_ids).order_by('-year')
            # Показываем столбец "Год" только если выбрано больше одного года
            show_year_column = len(selected_years) > 1
            # Сохраняем выбранные годы в сессию
            request.session['selected_year_ids'] = year_ids
        else:
            # Если годы не выбраны, используем все годы без столбца "Год"
            selected_years = years
            show_year_column = False
            # Очищаем сессию
            if 'selected_year_ids' in request.session:
                del request.session['selected_year_ids']
    else: # Обработка GET-запроса
        # Проверяем, есть ли сохраненный параметр show_ter_status в сессии
        show_ter_status = request.session.get('show_ter_status', False)
        
        # Проверяем, есть ли сохраненные годы в сессии
        if 'selected_year_ids' in request.session:
            year_ids = request.session.get('selected_year_ids')
            selected_years = Year.objects.filter(id__in=year_ids).order_by('-year')
            show_year_column = len(selected_years) > 1
        else:
            # По умолчанию используем все годы без столбца "Год"
            selected_years = years
            show_year_column = False
    
    # --- DEBUG START ---
    print(f"Final selected_years: {[y.year for y in selected_years]}")
    print(f"Final show_year_column: {show_year_column}")
    print(f"Final show_ter_status: {show_ter_status}")
    print(f"Session selected_year_ids after processing: {request.session.get('selected_year_ids')}")
    # --- DEBUG END ---
    
    # Если есть выбранные годы, рассчитываем статистику
    if selected_years:
        # Если не нужно показывать столбец "Год", агрегируем данные по всем годам
        if not show_year_column:
            # Рассчитываем статистику для всех лет вместе
            ter_admin_stats = {} # Словарь для агрегации данных по ТУ/ДО
            
            # Получаем отчеты для всех выбранных годов
            reports = Report.objects.filter(year__in=selected_years, is_published=True)
            
            # Фильтр школьных отчетов
            schools_reports_filter = {
                'report__in': reports,
                'status__in': ['A', 'D'] if show_ter_status else ['D']  # Учитываем отчеты на согласовании, если включен фильтр
            }
            
            # Если пользователь - представитель ТУ/ДО, добавляем фильтрацию по его ТУ/ДО
            if is_ter_admin_rep:
                ter_admin_ids = ter_admins.values_list('id', flat=True)
                schools_reports_filter['school__ter_admin_id__in'] = ter_admin_ids
            
            # Получаем отчеты школ с применением фильтров
            schools_reports = SchoolReport.objects.filter(**schools_reports_filter)
                
            # Подсчитываем отчеты по зонам для каждого ТУ/ДО
            for report in schools_reports:
                # Пропускаем отчеты без школы или с архивной школой
                if not report.school or report.school.is_archived:
                    continue
                
                ter_admin_id = report.school.ter_admin_id
                
                # Инициализируем словарь для ТУ/ДО, если его еще нет
                if ter_admin_id not in ter_admin_stats:
                    ter_admin = TerAdmin.objects.get(id=ter_admin_id)
                    ter_admin_stats[ter_admin_id] = {
                        'ter_admin_id': ter_admin_id,
                        'ter_admin_name': ter_admin.name,
                        'red_zone': 0,
                        'yellow_zone': 0,
                        'green_zone': 0,
                        'total': 0
                    }
                
                # Увеличиваем счетчик для соответствующей зоны
                if report.zone == 'R':
                    ter_admin_stats[ter_admin_id]['red_zone'] += 1
                elif report.zone == 'Y':
                    ter_admin_stats[ter_admin_id]['yellow_zone'] += 1
                elif report.zone == 'G':
                    ter_admin_stats[ter_admin_id]['green_zone'] += 1
                
                # Увеличиваем общий счетчик отчетов для ТУ/ДО
                ter_admin_stats[ter_admin_id]['total'] += 1
                
            # Рассчитываем проценты для каждого ТУ/ДО
            for ter_admin_id in ter_admin_stats:
                item = ter_admin_stats[ter_admin_id]
                if item['total'] > 0:
                    item['red_percent'] = f"{(item['red_zone'] / item['total']) * 100:.1f}".replace(',', '.')
                    item['yellow_percent'] = f"{(item['yellow_zone'] / item['total']) * 100:.1f}".replace(',', '.')
                    item['green_percent'] = f"{(item['green_zone'] / item['total']) * 100:.1f}".replace(',', '.')
                else:
                    item['red_percent'] = "0.0"
                    item['yellow_percent'] = "0.0"
                    item['green_percent'] = "0.0"
                    
            # Преобразуем словарь в список и сортируем по имени ТУ/ДО
            stats = list(ter_admin_stats.values())
            stats.sort(key=lambda x: x['ter_admin_name'])
            
            # Для отладки
            print(f"DEBUG - One year or all years: {len(stats)} items in stats")
            
        else:
            # Если нужно показывать столбец "Год", собираем данные по каждому году отдельно
            year_stats_list = []
            
            for year in selected_years:
                # Получаем отчеты для выбранного года
                reports = Report.objects.filter(year=year, is_published=True)
                
                # Фильтр школьных отчетов
                schools_reports_filter = {
                    'report__in': reports,
                    'status__in': ['A', 'D'] if show_ter_status else ['D']  # Учитываем отчеты на согласовании, если включен фильтр
                }
                
                # Если пользователь - представитель ТУ/ДО, добавляем фильтрацию по его ТУ/ДО
                if is_ter_admin_rep:
                    ter_admin_ids = ter_admins.values_list('id', flat=True)
                    schools_reports_filter['school__ter_admin_id__in'] = ter_admin_ids
                
                # Получаем отчеты школ с применением фильтров
                schools_reports = SchoolReport.objects.filter(**schools_reports_filter)
                
                # Словарь для текущего года
                year_stats = {}
                
                # Подсчитываем отчеты по зонам для каждого ТУ/ДО
                for report in schools_reports:
                    # Пропускаем отчеты без школы или с архивной школой
                    if not report.school or report.school.is_archived:
                        continue
                    
                    ter_admin_id = report.school.ter_admin_id
                    
                    # Инициализируем словарь для ТУ/ДО, если его еще нет
                    if ter_admin_id not in year_stats:
                        ter_admin = TerAdmin.objects.get(id=ter_admin_id)
                        year_stats[ter_admin_id] = {
                            'year': year.year,
                            'ter_admin_id': ter_admin_id,
                            'ter_admin_name': ter_admin.name,
                            'red_zone': 0,
                            'yellow_zone': 0,
                            'green_zone': 0,
                            'total': 0
                        }
                    
                    # Увеличиваем счетчик для соответствующей зоны
                    if report.zone == 'R':
                        year_stats[ter_admin_id]['red_zone'] += 1
                    elif report.zone == 'Y':
                        year_stats[ter_admin_id]['yellow_zone'] += 1
                    elif report.zone == 'G':
                        year_stats[ter_admin_id]['green_zone'] += 1
                    
                    # Увеличиваем общий счетчик отчетов для ТУ/ДО
                    year_stats[ter_admin_id]['total'] += 1
                
                # Рассчитываем проценты для каждого ТУ/ДО за текущий год
                for ter_admin_id in year_stats:
                    item = year_stats[ter_admin_id]
                    if item['total'] > 0:
                        item['red_percent'] = f"{(item['red_zone'] / item['total']) * 100:.1f}".replace(',', '.')
                        item['yellow_percent'] = f"{(item['yellow_zone'] / item['total']) * 100:.1f}".replace(',', '.')
                        item['green_percent'] = f"{(item['green_zone'] / item['total']) * 100:.1f}".replace(',', '.')
                    else:
                        item['red_percent'] = "0.0"
                        item['yellow_percent'] = "0.0"
                        item['green_percent'] = "0.0"
                
                # Преобразуем словарь в список и сортируем по имени ТУ/ДО
                current_year_stats = list(year_stats.values())
                current_year_stats.sort(key=lambda x: x['ter_admin_name'])
                
                # Добавляем статистику за текущий год в общий список
                year_stats_list.extend(current_year_stats)
            
            # Для отладки
            print(f"DEBUG - Multiple years: {len(year_stats_list)} total items across years")
            
            # Сохраняем несгруппированную статистику для использования в Excel
            stats_for_excel = year_stats_list
            
            # Группируем данные по ТУ/ДО для отображения в шаблоне
            grouped_stats = {}
            for item in year_stats_list:
                ter_admin_name = item['ter_admin_name']
                if ter_admin_name not in grouped_stats:
                    grouped_stats[ter_admin_name] = []
                grouped_stats[ter_admin_name].append(item)
            
            # Преобразуем сгруппированные данные в формат для шаблона
            stats_grouped = []
            for ter_admin_name in sorted(grouped_stats.keys()):
                # Сортируем элементы группы по году (по убыванию)
                items = sorted(grouped_stats[ter_admin_name], key=lambda x: x.get('year', 0), reverse=True)
                stats_grouped.append({
                    'ter_admin_name': ter_admin_name,
                    'years': items,
                    'rowspan': len(items)
                })
            
            # Используем сгруппированную статистику для отображения
            stats = stats_grouped
            
            # Для отладки
            print(f"DEBUG - After grouping: {len(stats)} ТУ/ДО groups")
    
    # Если пользователь - представитель ТУ/ДО, получаем список идентификаторов его ТУ/ДО
    ter_admin_ids = ter_admins.values_list('id', flat=True) if is_ter_admin_rep else None
    
    # Получаем статистику по разделам и кластерам
    section_stats = calculate_section_stats(selected_years, show_year_column, show_ter_status, ter_admin_ids)
    cluster_stats = calculate_cluster_stats(selected_years, show_year_column, show_ter_status, ter_admin_ids)
    
    # Для отладки
    print(f"DEBUG - section_stats: {len(section_stats) if section_stats else 0} items")
    print(f"DEBUG - cluster_stats: {len(cluster_stats) if cluster_stats else 0} items")
    
    # Подготавливаем контекст для шаблона
    context = {
        'filter': filter_context,
        'stats': stats,
        'selected_years': selected_years,
        'selected_year_ids': [str(year.id) for year in selected_years] if selected_years else [],
        'show_year_column': show_year_column,  # Флаг, показывающий, нужен ли столбец с годом
        'show_ter_status': show_ter_status  # Флаг, показывающий, нужен ли столбец с статусом ТУ/ДО
    }
    
    # Добавляем статистику по разделам и кластерам
    if section_stats:
        context['section_stats'] = section_stats
    
    if cluster_stats:
        context['cluster_stats'] = cluster_stats
        
    # Генерируем статистику по индикаторам (показателям)
    if selected_years:
        # Собираем статистику по показателям для каждого раздела
        indicator_stats = []
        
        # Получаем все разделы для выбранных годов, сгруппированные по имени
        section_names = Section.objects.filter(
            report__year__in=selected_years
        ).values_list('name', 'number').distinct()
        
        # Создаем словарь для быстрого доступа
        section_dict = {name: number for name, number in section_names}
        
        # Получаем все отчеты школ для выбранных годов
        all_school_reports_filter = {
            'report__year__in': selected_years,
            'status__in': ['A', 'D'] if show_ter_status else ['D']  # Учитываем отчеты на согласовании
        }
        
        # Если пользователь - представитель ТУ/ДО, добавляем фильтрацию по его ТУ/ДО
        if is_ter_admin_rep:
            ter_admin_ids = ter_admins.values_list('id', flat=True)
            all_school_reports_filter['school__ter_admin_id__in'] = ter_admin_ids
        
        all_school_reports = SchoolReport.objects.filter(**all_school_reports_filter).exclude(
            school=None
        ).exclude(
            school__is_archived=True
        ).select_related(
            'school__ter_admin',
            'school__closter',
            'report__year'
        ).prefetch_related(
            'sections',
            'answers'
        )
        
        # Для каждого раздела собираем данные по показателям
        for section_name, section_number in section_names:
            # Находим все разделы с указанным именем
            same_name_sections = Section.objects.filter(
                name=section_name,
                report__year__in=selected_years
            )
            
            # Получаем все поля (показатели) для текущего раздела
            fields = Field.objects.filter(
                sections__in=same_name_sections
            ).distinct().values('id', 'name', 'number')
            
            section_indicators = []
            
            # Для каждого поля (показателя) собираем статистику по ТУ/ДО
            for field in fields:
                field_id = field['id']
                
                # Словарь для сбора статистики по ТУ/ДО
                ter_admin_stats = {}
                
                # Получаем все ответы по этому показателю
                answers = Answer.objects.filter(
                    question_id=field_id,
                    s_report__in=all_school_reports
                ).select_related('s_report__school__ter_admin')
                
                # Создаем уникальный набор комбинаций [отчет, ТУ/ДО] для избежания дублей
                processed_reports = set()
                
                # Собираем статистику по зонам для каждого ТУ/ДО
                for answer in answers:
                    s_report = answer.s_report
                    school = s_report.school
                    
                    if not school or not school.ter_admin:
                        continue
                    
                    # Создаем уникальный ключ для комбинации отчета и ТУ/ДО
                    key = (s_report.id, school.ter_admin.id)
                    
                    # Пропускаем, если уже обработали этот отчет для этого ТУ/ДО
                    if key in processed_reports:
                        continue
                    
                    processed_reports.add(key)
                    
                    ter_admin_name = school.ter_admin.name
                    
                    # Инициализируем счетчики для ТУ/ДО, если их еще нет
                    if ter_admin_name not in ter_admin_stats:
                        ter_admin_stats[ter_admin_name] = {
                            'ter_admin_name': ter_admin_name,
                            'red_zone': 0,
                            'yellow_zone': 0,
                            'green_zone': 0,
                            'total': 0
                        }
                    
                    # Увеличиваем счетчик для соответствующей зоны
                    if answer.zone == 'R':
                        ter_admin_stats[ter_admin_name]['red_zone'] += 1
                    elif answer.zone == 'Y':
                        ter_admin_stats[ter_admin_name]['yellow_zone'] += 1
                    elif answer.zone == 'G':
                        ter_admin_stats[ter_admin_name]['green_zone'] += 1
                    
                    # Увеличиваем общий счетчик
                    ter_admin_stats[ter_admin_name]['total'] += 1
                
                # Рассчитываем проценты для каждого ТУ/ДО
                for ter_admin_name in ter_admin_stats:
                    item = ter_admin_stats[ter_admin_name]
                    if item['total'] > 0:
                        item['red_percent'] = f"{(item['red_zone'] / item['total']) * 100:.1f}".replace(',', '.')
                        item['yellow_percent'] = f"{(item['yellow_zone'] / item['total']) * 100:.1f}".replace(',', '.')
                        item['green_percent'] = f"{(item['green_zone'] / item['total']) * 100:.1f}".replace(',', '.')
                    else:
                        item['red_percent'] = "0.0"
                        item['yellow_percent'] = "0.0"
                        item['green_percent'] = "0.0"
                
                # Преобразуем в список и сортируем по имени ТУ/ДО
                ter_admin_list = list(ter_admin_stats.values())
                ter_admin_list.sort(key=lambda x: x['ter_admin_name'])
                
                # Если есть данные для этого поля, добавляем его в список
                if ter_admin_list:
                    section_indicators.append({
                        'field_id': field_id,
                        'field_name': field['name'],
                        'field_number': field['number'],
                        'ter_admin_stats': ter_admin_list
                    })
            
            # Сортируем показатели по номеру
            section_indicators.sort(key=lambda x: [int(n) for n in str(x['field_number']).split('.')] if x['field_number'] else [0])
            
            # Если есть показатели для этого раздела, добавляем его в список
            if section_indicators:
                indicator_stats.append({
                    'section_number': section_number,
                    'section_name': section_name,
                    'indicators': section_indicators
                })
        
        # Сортируем разделы по номеру
        indicator_stats.sort(key=lambda x: [int(n) for n in str(x['section_number']).split('.')] if x['section_number'] else [0])
        
        # Добавляем данные в контекст
        context['indicator_stats'] = indicator_stats
    
    # Обрабатываем запрос на скачивание Excel-файла
    if request.method == 'POST' and 'download' in request.POST:
        try:
            # Для выгрузки в Excel используем подходящие данные
            if show_year_column:
                # Для нескольких лет используем несгруппированный список
                if 'stats_for_excel' in locals():
                    excel_data = stats_for_excel
                else:
                    # Если stats_for_excel не определен, но stats сгруппирован
                    if stats and isinstance(stats, list) and stats and isinstance(stats[0], dict) and 'years' in stats[0]:
                        # Разгруппировываем данные
                        flat_stats = []
                        for group in stats:
                            flat_stats.extend(group.get('years', []))
                        excel_data = flat_stats
                    else:
                        # Если статистика не сгруппирована
                        excel_data = stats
            else:
                # Для одного года просто используем stats
                excel_data = stats
            
            # Убедимся, что excel_data не пустой
            if not excel_data:
                excel_data = []
                print("WARNING: Empty excel_data!")
            
            # Дополнительная проверка структуры excel_data для отладки
            print(f"DEBUG - excel_data type: {type(excel_data)}")
            if excel_data and isinstance(excel_data, list):
                print(f"DEBUG - First item in excel_data: {excel_data[0].keys() if isinstance(excel_data[0], dict) else 'Not a dict'}")
            
            # Создаем Excel-файл
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="report_distribution.xlsx"'
            
            # Генерируем Excel с распределением зон
            print(f"DEBUG - Before generate_zone_distribution_excel")
            wb = generate_zone_distribution_excel(
                selected_years, 
                excel_data, 
                show_year_column, 
                section_stats, 
                cluster_stats, 
                indicator_stats,
                show_ter_status,
                ter_admin_ids
            )
            print(f"DEBUG - After generate_zone_distribution_excel")
            return wb
        except Exception as e:
            # Логируем ошибку
            print(f"ERROR when generating Excel: {str(e)}")
            import traceback
            print(traceback.format_exc())
            
            # Возвращаем сообщение об ошибке
            messages.error(request, f"Ошибка при формировании Excel-отчета: {str(e)}")
            # Возвращаем страницу без скачивания файла
            return render(request, 'dashboards/answers_distribution_report.html', context)

    # Возвращаем шаблон с контекстом
    return render(request, 'dashboards/answers_distribution_report.html', context)

def generate_zone_distribution_excel(years, stats, show_year_column=False, section_stats=None, cluster_stats=None, indicator_stats=None, show_ter_status=False, ter_admin_ids=None):
    """
    Генерирует Excel-файл с распределением зон по ТУ/ДО, разделам и кластерам,
    а также с детальной информацией по школам и показателям
    """
    # Проверяем и исправляем входные данные
    stats = stats or []
    section_stats = section_stats or []
    cluster_stats = cluster_stats or []
    
    # Если stats - это список словарей с ключом 'years', разгруппируем его
    if show_year_column and stats and isinstance(stats, list) and stats and isinstance(stats[0], dict) and 'years' in stats[0]:
        flat_stats = []
        for group in stats:
            flat_stats.extend(group.get('years', []))
        stats = flat_stats
    
    # Создаем Excel-файл
    output = BytesIO()
    workbook = Workbook()
    workbook.active.title = "Визуализация"
    
    # Стили для ячеек (определяем один раз для повторного использования)
    styles = {
        'red_fill': PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'),
        'yellow_fill': PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),
        'green_fill': PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid'),
        'bold_font': Font(bold=True)
    }
    
    # ========== ВКЛАДКА "ВИЗУАЛИЗАЦИЯ" ==========
    _create_visualization_sheet(workbook.active, stats, show_year_column, styles, section_stats, cluster_stats)
    
    # ========== ВКЛАДКА "ОБЩИЙ СВОД" ==========
    general_sheet = workbook.create_sheet(title="Общий свод")
    
    # Оптимизированный запрос для получения всех отчетов школ с предзагрузкой связанных данных
    all_school_reports = _get_all_school_reports(years, show_ter_status, ter_admin_ids)
    
    # Создаем общий свод
    _create_general_sheet(general_sheet, all_school_reports, years, styles)
    
    # ========== ВКЛАДКИ ДЛЯ КАЖДОГО РАЗДЕЛА ==========
    _create_section_sheets(workbook, all_school_reports, years, styles)
    
    # Сохраняем файл
    workbook.save(output)
    output.seek(0)
    
    # Генерируем HTTP-ответ
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    years_str = "-".join([str(year.year) for year in years])
    response['Content-Disposition'] = f'attachment; filename="zone_distribution_{years_str}.xlsx"'
    
    return response

# Вспомогательные функции для generate_zone_distribution_excel
def _create_visualization_sheet(sheet, stats, show_year_column, styles, section_stats, cluster_stats):
    """Создает вкладку визуализации в Excel-файле"""
    # Заголовок
    row_num = 1
    cell = sheet.cell(row=row_num, column=1, value="Распределение отчетов по зонам")
    cell.font = styles['bold_font']
    row_num += 2
    
    # Заголовки столбцов
    headers = ["ТУ/ДО"]
    if show_year_column:
        headers.append("Год")
    headers.extend(["Красная зона", "Жёлтая зона", "Зелёная зона", "Всего отчетов"])
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=row_num, column=col_num, value=header)
        cell.font = styles['bold_font']
    row_num += 1
    
    try:
        if not stats:
            sheet.cell(row=row_num, column=1, value="Нет данных для отображения")
        elif show_year_column:
            row_num = _add_multi_year_stats(sheet, stats, row_num, styles)
        else:
            row_num = _add_single_year_stats(sheet, stats, row_num, styles)
            
        # Добавляем таблицы по разделам и кластерам
        if section_stats:
            row_num = _add_section_stats(sheet, section_stats, show_year_column, row_num, styles)
        
        if cluster_stats:
            row_num = _add_cluster_stats(sheet, cluster_stats, show_year_column, row_num, styles)
    except Exception as e:
        sheet.cell(row=row_num, column=1, value=f"Ошибка обработки данных: {str(e)}")
    
    # Автоподбор ширины столбцов
    _adjust_column_width(sheet)

def _add_multi_year_stats(sheet, stats, row_num, styles):
    """Добавляет статистику за несколько лет в таблицу"""
    if not all(isinstance(item, dict) and 'ter_admin_name' in item for item in stats):
        sheet.cell(row=row_num, column=1, value="Ошибка в структуре данных")
        return row_num + 1
        
    # Группируем данные по ter_admin_name
    ter_admin_groups = {}
    for item in stats:
        ter_admin_name = item.get('ter_admin_name', 'Неизвестно')
        if ter_admin_name not in ter_admin_groups:
            ter_admin_groups[ter_admin_name] = []
        ter_admin_groups[ter_admin_name].append(item)
    
    # Отображаем данные по группам с объединением ячеек
    for ter_admin_name in sorted(ter_admin_groups.keys()):
        group_items = ter_admin_groups[ter_admin_name]
        start_row = row_num
        
        # Сортируем элементы группы по году
        group_items.sort(key=lambda x: x.get('year', 0), reverse=True)
        
        # Проходим по всем годам для текущего ТУ/ДО
        for item in group_items:
            col = 1
            
            # Название ТУ/ДО (добавляем только в первую строку группы)
            if row_num == start_row:
                sheet.cell(row=row_num, column=col, value=ter_admin_name)
            col += 1
            
            # Год
            sheet.cell(row=row_num, column=col, value=item.get('year', ''))
            col += 1
            
            # Данные по зонам
            row_num = _add_zone_cells(sheet, row_num, col, item, styles)
        
        # Объединяем ячейки с названием ТУ/ДО, если группа содержит более одной записи
        if len(group_items) > 1:
            try:
                sheet.merge_cells(start_row=start_row, start_column=1, end_row=row_num-1, end_column=1)
            except Exception:
                pass
    
    return row_num

def _add_single_year_stats(sheet, stats, row_num, styles):
    """Добавляет статистику за один год в таблицу"""
    # Сортируем статистику по имени ТУ/ДО
    sorted_stats = sorted(stats, key=lambda x: x.get('ter_admin_name', ''))
    for item in sorted_stats:
        col = 1
        
        # Название ТУ/ДО
        sheet.cell(row=row_num, column=col, value=item.get('ter_admin_name', 'Неизвестно'))
        col += 1
        
        # Данные по зонам
        row_num = _add_zone_cells(sheet, row_num, col, item, styles)
    
    return row_num

def _add_zone_cells(sheet, row_num, col, item, styles):
    """Добавляет ячейки с данными по зонам в таблицу"""
    # Красная зона
    red_zone_value = item.get('red_zone', 0)
    cell_value = "-" if red_zone_value == 0 else red_zone_value
    cell = sheet.cell(row=row_num, column=col, value=cell_value)
    if red_zone_value > 0:
        cell.fill = styles['red_fill']
    col += 1
    
    # Жёлтая зона
    yellow_zone_value = item.get('yellow_zone', 0)
    cell_value = "-" if yellow_zone_value == 0 else yellow_zone_value
    cell = sheet.cell(row=row_num, column=col, value=cell_value)
    if yellow_zone_value > 0:
        cell.fill = styles['yellow_fill']
    col += 1
    
    # Зелёная зона
    green_zone_value = item.get('green_zone', 0)
    cell_value = "-" if green_zone_value == 0 else green_zone_value
    cell = sheet.cell(row=row_num, column=col, value=cell_value)
    if green_zone_value > 0:
        cell.fill = styles['green_fill']
    col += 1
    
    # Всего отчетов
    total_value = item.get('total', 0)
    cell_value = "-" if total_value == 0 else total_value
    sheet.cell(row=row_num, column=col, value=cell_value)
    
    return row_num + 1

def _add_section_stats(sheet, section_stats, show_year_column, row_num, styles):
    """Добавляет статистику по разделам в таблицу"""
    row_num += 2
    sheet.cell(row=row_num, column=1, value="Распределение отчетов по разделам").font = styles['bold_font']
    row_num += 2
    
    # Заголовки столбцов
    headers = ["Раздел"]
    if show_year_column:
        headers.append("Год")
    headers.extend(["Красная зона", "Жёлтая зона", "Зелёная зона", "Всего отчетов"])
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=row_num, column=col_num, value=header)
        cell.font = styles['bold_font']
    row_num += 1
    
    # Добавляем данные по разделам
    if show_year_column and section_stats and isinstance(section_stats[0], dict) and 'years' in section_stats[0]:
        # Группируем данные по названию раздела для нескольких лет
        for section_data in section_stats:
            section_name = section_data['section_name']
            start_row = row_num
            
            # Сортируем годы
            years_data = section_data['years']
            years_data.sort(key=lambda x: x.get('year', 0))
            
            # Добавляем данные для каждого года
            for item in years_data:
                col = 1
                
                # Название раздела (только для первой строки)
                if row_num == start_row:
                    sheet.cell(row=row_num, column=col, value=section_name)
                col += 1
                
                # Год
                sheet.cell(row=row_num, column=col, value=item.get('year', ''))
                col += 1
                
                # Данные по зонам
                row_num = _add_zone_cells(sheet, row_num, col, item, styles)
            
            # Объединяем ячейки с названием раздела
            if len(years_data) > 1:
                try:
                    sheet.merge_cells(start_row=start_row, start_column=1, end_row=row_num-1, end_column=1)
                except Exception:
                    pass
    else:
        # Данные за один год
        for item in section_stats:
            col = 1
            
            # Название раздела
            sheet.cell(row=row_num, column=col, value=item['section_name'])
            col += 1
            
            # Год (если нужно)
            if show_year_column:
                sheet.cell(row=row_num, column=col, value=item.get('year', ''))
                col += 1
            
            # Данные по зонам
            row_num = _add_zone_cells(sheet, row_num, col, item, styles)
    
    return row_num

def _add_cluster_stats(sheet, cluster_stats, show_year_column, row_num, styles):
    """Добавляет статистику по кластерам в таблицу"""
    row_num += 2
    sheet.cell(row=row_num, column=1, value="Распределение отчетов по кластерам").font = styles['bold_font']
    row_num += 2
    
    # Заголовки столбцов
    headers = ["Кластер"]
    if show_year_column:
        headers.append("Год")
    headers.extend(["Красная зона", "Жёлтая зона", "Зелёная зона", "Всего отчетов"])
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=row_num, column=col_num, value=header)
        cell.font = styles['bold_font']
    row_num += 1
    
    # Добавляем данные по кластерам
    if show_year_column and cluster_stats and isinstance(cluster_stats[0], dict) and 'years' in cluster_stats[0]:
        # Группируем данные по названию кластера для нескольких лет
        for cluster_data in cluster_stats:
            cluster_name = cluster_data['cluster_name']
            start_row = row_num
            
            # Сортируем годы
            years_data = cluster_data['years']
            years_data.sort(key=lambda x: x.get('year', 0))
            
            # Добавляем данные для каждого года
            for item in years_data:
                col = 1
                
                # Название кластера (только для первой строки)
                if row_num == start_row:
                    sheet.cell(row=row_num, column=col, value=cluster_name)
                col += 1
                
                # Год
                sheet.cell(row=row_num, column=col, value=item.get('year', ''))
                col += 1
                
                # Данные по зонам
                row_num = _add_zone_cells(sheet, row_num, col, item, styles)
            
            # Объединяем ячейки с названием кластера
            if len(years_data) > 1:
                try:
                    sheet.merge_cells(start_row=start_row, start_column=1, end_row=row_num-1, end_column=1)
                except Exception:
                    pass
    else:
        # Данные за один год
        for item in cluster_stats:
            col = 1
            
            # Название кластера
            sheet.cell(row=row_num, column=col, value=item['cluster_name'])
            col += 1
            
            # Год (если нужно)
            if show_year_column:
                sheet.cell(row=row_num, column=col, value=item.get('year', ''))
                col += 1
            
            # Данные по зонам
            row_num = _add_zone_cells(sheet, row_num, col, item, styles)
    
    return row_num

def _get_all_school_reports(years, show_ter_status, ter_admin_ids):
    """Получает все отчеты школ с предзагрузкой связанных данных"""
    filter_params = {
        'report__year__in': years,
        'status__in': ['A', 'D'] if show_ter_status else ['D']  # Учитываем отчеты на согласовании
    }
    
    # Если указаны ТУ/ДО, добавляем их в фильтр
    if ter_admin_ids:
        filter_params['school__ter_admin_id__in'] = ter_admin_ids
    
    return SchoolReport.objects.filter(**filter_params).exclude(
        school=None
    ).exclude(
        school__is_archived=True
    ).select_related(
        'school__ter_admin',
        'school__closter',
        'report__year'
    ).prefetch_related(
        Prefetch('sections', queryset=SectionSreport.objects.select_related('section')),
        Prefetch('answers', queryset=Answer.objects.select_related('question'))
    )

def _create_general_sheet(sheet, all_school_reports, years, styles):
    """Создает вкладку общего свода в Excel-файле"""
    # Заголовок
    row_num = 1
    sheet.cell(row=row_num, column=1, value="Общий свод по школам").font = styles['bold_font']
    row_num += 2
    
    # Получаем все разделы для выбранных годов для "Общего свода"
    sections = Section.objects.filter(
        report__year__in=years
    ).distinct('number').order_by('number')
    
    # Заголовки столбцов
    headers = ["Школа", "ТУ/ДО", "Уровень образования", "Кластер", "Итого баллов"]
    headers.extend([f"Раздел {s.number}" for s in sections])
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=row_num, column=col_num, value=header)
        cell.font = styles['bold_font']
    row_num += 1
    
    # Словарь для хранения уровней образования
    ed_levels = {
        'A': "1 — 11 классы",
        'M': "1 — 9 классы",
        'S': "1 — 4 классы",
        'MG': "5 — 11 классы",
        'G': "10 — 11 классы",
    }
    
    # Группируем отчеты по школам сразу в словаре, а не в цикле
    school_data = defaultdict(lambda: {
        'sections': defaultdict(float),
        'total_points': 0,
        'year': None,
        'zone': None
    })
    
    # Обрабатываем отчеты и заполняем словарь
    for s_report in all_school_reports:
        school = s_report.school
        if not school:
            continue
            
        school_id = school.id
        year = s_report.report.year
        key = (school_id, year.id)
        
        if 'school_name' not in school_data[key]:
            school_data[key].update({
                'school_name': school.name,
                'ter_admin_name': school.ter_admin.name if school.ter_admin else "-",
                'ed_level': ed_levels.get(school.ed_level, "Н/Д"),
                'closter_name': school.closter.name if school.closter else "-",
                'year': year.year,
                'zone': s_report.zone
            })
        
        # Добавляем данные по секциям за один проход
        for section in s_report.sections.all():
            section_number = section.section.number
            school_data[key]['sections'][section_number] = section.points
            school_data[key]['total_points'] += section.points
    
    # Сортируем данные школ
    sorted_school_data = sorted(
        school_data.values(),
        key=lambda x: (x['ter_admin_name'], x['school_name'], x.get('year', 0))
    )
    
    # Заполняем данные школ одним проходом
    for data in sorted_school_data:
        col = 1
        
        # Базовая информация о школе
        sheet.cell(row=row_num, column=col, value=data['school_name']); col += 1
        sheet.cell(row=row_num, column=col, value=data['ter_admin_name']); col += 1
        sheet.cell(row=row_num, column=col, value=data['ed_level']); col += 1
        sheet.cell(row=row_num, column=col, value=data['closter_name']); col += 1
        
        # Итого баллов с цветом зоны
        cell = sheet.cell(row=row_num, column=col, value=data['total_points'])
        zone = data.get('zone')
        if zone == 'R':
            cell.fill = styles['red_fill']
        elif zone == 'Y':
            cell.fill = styles['yellow_fill']
        elif zone == 'G':
            cell.fill = styles['green_fill']
        col += 1
        
        # Баллы по разделам
        for section in sections:
            section_points = data['sections'].get(section.number, "-")
            sheet.cell(row=row_num, column=col, value=section_points)
            col += 1
        
        row_num += 1
    
    # Автоподбор ширины столбцов
    _adjust_column_width(sheet)

def _create_section_sheets(workbook, all_school_reports, years, styles):
    """Создает вкладки для каждого раздела в Excel-файле"""
    # Получаем все разделы для выбранных годов
    all_sections = Section.objects.filter(
        report__year__in=years
    ).values('name', 'number').distinct('number').order_by('number')
    
    # Словарь для уровней образования
    ed_levels = {
        'A': "1 — 11 классы",
        'M': "1 — 9 классы",
        'S': "1 — 4 классы",
        'MG': "5 — 11 классы",
        'G': "10 — 11 классы",
    }
    
    # Предзагружаем все поля для всех разделов
    all_fields = {}
    section_numbers = [section['number'] for section in all_sections]
    for section_number in section_numbers:
        fields = Field.objects.filter(
            sections__number=section_number,
            sections__report__year__in=years
        ).distinct()
        # Применяем правильную числовую сортировку по номерам полей
        fields = sorted(fields, key=lambda x: [int(n) for n in str(x.number).split('.')])
        all_fields[section_number] = list(fields)
    
    # Создаем отдельный лист для каждого раздела
    for section in all_sections:
        section_name = section['name']
        section_number = section['number']
        
        # Создаем имя листа без недопустимых символов (макс. 31 символ)
        safe_sheet_name = f"Раздел {section_number}"
        if len(safe_sheet_name) > 31:
            safe_sheet_name = safe_sheet_name[:28] + "..."
        
        section_sheet = workbook.create_sheet(title=safe_sheet_name)
        
        # Заголовок
        row_num = 1
        section_sheet.cell(row=row_num, column=1, value=f"Раздел {section_number}: {section_name}").font = styles['bold_font']
        row_num += 2
        
        # Получаем поля (показатели) для текущего раздела
        fields = all_fields.get(section_number, [])
        
        # Заголовки столбцов: базовые + показатели
        headers = ["Школа", "ТУ/ДО", "Уровень образования", "Кластер"]
        field_headers = [f"{field.number}. {field.name}" for field in fields]
        headers.extend(field_headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = section_sheet.cell(row=row_num, column=col_num, value=header)
            cell.font = styles['bold_font']
        row_num += 1
        
        # Оптимизированная обработка данных по школам
        school_data = {}
        
        # Создаем словарь с ID полей для быстрого поиска
        field_ids = {field.id: field for field in fields}
        
        # Находим и группируем отчеты и ответы для этого раздела
        for s_report in all_school_reports:
            school = s_report.school
            if not school:
                continue
                
            # Пропускаем, если нет секции для этого раздела
            has_section = False
            for section_sreport in s_report.sections.all():
                if section_sreport.section.number == section_number:
                    has_section = True
                    break
                    
            if not has_section:
                continue
                
            school_id = school.id
            
            if school_id not in school_data:
                school_data[school_id] = {
                    'school_name': school.name,
                    'ter_admin_name': school.ter_admin.name if school.ter_admin else "-",
                    'ed_level': ed_levels.get(school.ed_level, "Н/Д"),
                    'closter_name': school.closter.name if school.closter else "-",
                    'field_data': {}
                }
            
            # Обрабатываем ответы на показатели этого раздела
            # Используем предзагруженные ответы
            for answer in s_report.answers.all():
                field_id = answer.question_id
                
                # Проверяем, что ответ относится к этому разделу
                if field_id not in field_ids:
                    continue
                    
                field_data = {
                    'points': answer.points,
                    'zone': answer.zone,
                    'year': s_report.report.year.year
                }
                
                # Обновляем данные, только если ответ новее имеющегося
                if field_id in school_data[school_id]['field_data']:
                    if field_data['year'] > school_data[school_id]['field_data'][field_id].get('year', 0):
                        school_data[school_id]['field_data'][field_id] = field_data
                else:
                    school_data[school_id]['field_data'][field_id] = field_data
        
        # Сортируем и записываем данные школ
        sorted_school_data = sorted(
            school_data.values(),
            key=lambda x: (x['ter_admin_name'], x['school_name'])
        )
        
        for data in sorted_school_data:
            col = 1
            
            # Базовая информация о школе
            section_sheet.cell(row=row_num, column=col, value=data['school_name']); col += 1
            section_sheet.cell(row=row_num, column=col, value=data['ter_admin_name']); col += 1
            section_sheet.cell(row=row_num, column=col, value=data['ed_level']); col += 1
            section_sheet.cell(row=row_num, column=col, value=data['closter_name']); col += 1
            
            # Данные по показателям
            for field in fields:
                field_id = field.id
                field_data = data['field_data'].get(field_id, {})
                
                if field_data:
                    points = field_data.get('points', "-")
                    zone = field_data.get('zone', '')
                    
                    cell = section_sheet.cell(row=row_num, column=col, value=points)
                    
                    # Устанавливаем цвет ячейки в зависимости от зоны
                    if zone == 'R':
                        cell.fill = styles['red_fill']
                    elif zone == 'Y':
                        cell.fill = styles['yellow_fill']
                    elif zone == 'G':
                        cell.fill = styles['green_fill']
                else:
                    section_sheet.cell(row=row_num, column=col, value="-")
                
                col += 1
            
            row_num += 1
        
        # Автоподбор ширины столбцов
        _adjust_column_width(section_sheet)

def _adjust_column_width(sheet):
    """Регулирует ширину столбцов на листе Excel"""
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
                
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

# Вспомогательные функции для расчета статистики (добавить в конец файла)
def calculate_section_stats(selected_years, show_year_column=False, show_ter_status=False, ter_admin_ids=None):
    """
    Рассчитывает статистику по разделам для выбранных годов.
    """
    import re
    
    def normalize_section_name(name):
        """Нормализует название раздела: убирает лишние пробелы и приводит к lowercase для сравнения"""
        if not name:
            return ""
        # Убираем пробелы в начале и конце, заменяем множественные пробелы на одинарные
        normalized = re.sub(r'\s+', ' ', name.strip())
        return normalized.lower()
    
    section_stats = []
    section_stats_by_year = []
    
    # Получаем все разделы для выбранных годов
    all_sections = Section.objects.filter(
        report__year__in=selected_years
    ).values_list('name', flat=True).distinct()
    
    # Группируем разделы по нормализованному имени
    section_groups = {}
    for section_name in all_sections:
        normalized = normalize_section_name(section_name)
        if normalized not in section_groups:
            # Сохраняем первое встреченное оригинальное имя для отображения
            section_groups[normalized] = section_name
    
    for normalized_name, display_name in section_groups.items():
        if not show_year_column:
            # Находим все разделы, которые после нормализации имеют то же имя
            all_sections_for_year = Section.objects.filter(
                report__year__in=selected_years
            )
            same_name_sections = []
            for section in all_sections_for_year:
                if normalize_section_name(section.name) == normalized_name:
                    same_name_sections.append(section.id)
            
            # Преобразуем в QuerySet для дальнейшего использования
            from django.db.models import Q
            if not same_name_sections:
                continue
            
            same_name_sections_qs = Section.objects.filter(
                id__in=same_name_sections,
                report__year__in=selected_years
            ).values_list('id', flat=True)
            
            # Создаем базовый фильтр для SchoolReport
            school_reports_filter = {
                'report__year__in': selected_years, 
                'report__is_published': True,
                'status__in': ['A', 'D'] if show_ter_status else ['D']  # Учитываем отчеты на согласовании
            }
            
            # Если указаны ТУ/ДО, добавляем их в фильтр
            if ter_admin_ids:
                school_reports_filter['school__ter_admin_id__in'] = ter_admin_ids
                
            # Агрегированная статистика по разделам для всех годов
            school_reports = SchoolReport.objects.filter(**school_reports_filter).select_related('school')
            
            # Подсчитываем количество отчетов в каждой зоне для текущего раздела
            red_zone = 0
            yellow_zone = 0
            green_zone = 0
            total = 0
            
            # Для каждого отчета школы определяем зону данного раздела
            for sr in school_reports:
                # Пропускаем отчеты без школы или с архивной школой
                if not sr.school or sr.school.is_archived:
                    continue
                
                # Находим секцию отчета, соответствующую любому из разделов с тем же именем
                report_sections = sr.sections.filter(section__id__in=same_name_sections_qs)
                
                # Проверяем только одну секцию для каждого отчета (первую найденную)
                if report_sections.exists():
                    report_section = report_sections.first()
                    if report_section.zone == 'R':
                        red_zone += 1
                    elif report_section.zone == 'Y':
                        yellow_zone += 1
                    elif report_section.zone == 'G':
                        green_zone += 1
                    total += 1
            
            if total > 0:
                section_stats.append({
                    'section_name': display_name,
                    'red_zone': red_zone,
                    'yellow_zone': yellow_zone,
                    'green_zone': green_zone,
                    'total': total,
                    'red_percent': f"{(red_zone / total) * 100:.1f}".replace(',', '.'),
                    'yellow_percent': f"{(yellow_zone / total) * 100:.1f}".replace(',', '.'),
                    'green_percent': f"{(green_zone / total) * 100:.1f}".replace(',', '.')
                })
        else:
            # Статистика по разделам с разбивкой по годам
            section_years_stats = []
            
            for year in selected_years:
                # Находим все разделы, которые после нормализации имеют то же имя для текущего года
                all_sections_for_year = Section.objects.filter(
                    report__year=year
                )
                same_name_sections = []
                for section in all_sections_for_year:
                    if normalize_section_name(section.name) == normalized_name:
                        same_name_sections.append(section.id)
                
                # Если нет разделов с таким именем для этого года, пропускаем
                if not same_name_sections:
                    continue
                
                same_name_sections_qs = same_name_sections
                
                # Создаем базовый фильтр для SchoolReport
                school_reports_filter = {
                    'report__year': year, 
                    'report__is_published': True,
                    'status__in': ['A', 'D'] if show_ter_status else ['D']  # Учитываем отчеты на согласовании
                }
                
                # Если указаны ТУ/ДО, добавляем их в фильтр
                if ter_admin_ids:
                    school_reports_filter['school__ter_admin_id__in'] = ter_admin_ids
                
                school_reports = SchoolReport.objects.filter(**school_reports_filter).select_related('school')
                
                # Подсчитываем количество отчетов в каждой зоне для текущего раздела и года
                red_zone = 0
                yellow_zone = 0
                green_zone = 0
                total = 0
                
                # Для каждого отчета школы определяем зону данного раздела
                for sr in school_reports:
                    # Пропускаем отчеты без школы или с архивной школой
                    if not sr.school or sr.school.is_archived:
                        continue
                    
                    # Находим секцию отчета, соответствующую любому из разделов с тем же именем
                    report_sections = sr.sections.filter(section__id__in=same_name_sections_qs)
                    
                    # Проверяем только одну секцию для каждого отчета (первую найденную)
                    if report_sections.exists():
                        report_section = report_sections.first()
                        if report_section.zone == 'R':
                            red_zone += 1
                        elif report_section.zone == 'Y':
                            yellow_zone += 1
                        elif report_section.zone == 'G':
                            green_zone += 1
                        total += 1
                
                if total > 0:
                    # Данные по годам добавляем в отдельную структуру
                    section_years_stats.append({
                        'year': year.year,
                        'red_zone': red_zone,
                        'yellow_zone': yellow_zone,
                        'green_zone': green_zone,
                        'total': total,
                        'red_percent': f"{(red_zone / total) * 100:.1f}".replace(',', '.'),
                        'yellow_percent': f"{(yellow_zone / total) * 100:.1f}".replace(',', '.'),
                        'green_percent': f"{(green_zone / total) * 100:.1f}".replace(',', '.')
                    })
            
            if section_years_stats:
                # Группируем данные по разделам
                section_stats_by_year.append({
                    'section_name': display_name,
                    'years': section_years_stats,
                    'rowspan': len(section_years_stats)
                })
    
    # Выбираем правильную структуру данных в зависимости от режима отображения
    if show_year_column:
        return section_stats_by_year
    return section_stats

def calculate_cluster_stats(selected_years, show_year_column=False, show_ter_status=False, ter_admin_ids=None):
    """
    Рассчитывает статистику по кластерам для выбранных годов.
    """
    cluster_stats = []
    
    # Получаем все кластеры
    clusters = SchoolCloster.objects.all()
    
    for cluster in clusters:
        if not show_year_column:
            # Создаем базовый фильтр для SchoolReport
            school_reports_filter = {
                'report__year__in': selected_years, 
                'report__is_published': True,
                'status__in': ['A', 'D'] if show_ter_status else ['D'],  # Учитываем отчеты на согласовании
                'school__closter': cluster
            }
            
            # Если указаны ТУ/ДО, добавляем их в фильтр
            if ter_admin_ids:
                school_reports_filter['school__ter_admin_id__in'] = ter_admin_ids
            
            # Агрегированная статистика по кластерам для всех годов
            school_reports = SchoolReport.objects.filter(**school_reports_filter).select_related('school')
            
            # Подсчитываем количество отчетов в каждой зоне для текущего кластера
            red_zone = 0
            yellow_zone = 0
            green_zone = 0
            total = 0
            
            # Для каждого отчета школы определяем общую зону отчета
            for sr in school_reports:
                # Пропускаем отчеты без школы или с архивной школой
                if not sr.school or sr.school.is_archived:
                    continue
                
                if sr.zone == 'R':
                    red_zone += 1
                elif sr.zone == 'Y':
                    yellow_zone += 1
                elif sr.zone == 'G':
                    green_zone += 1
                total += 1
            
            if total > 0:
                cluster_stats.append({
                    'cluster_name': cluster.name,
                    'red_zone': red_zone,
                    'yellow_zone': yellow_zone,
                    'green_zone': green_zone,
                    'total': total,
                    'red_percent': f"{(red_zone / total) * 100:.1f}".replace(',', '.'),
                    'yellow_percent': f"{(yellow_zone / total) * 100:.1f}".replace(',', '.'),
                    'green_percent': f"{(green_zone / total) * 100:.1f}".replace(',', '.')
                })
        else:
            # Статистика по кластерам с разбивкой по годам
            cluster_years_stats = []
            
            for year in selected_years:
                # Создаем базовый фильтр для SchoolReport
                school_reports_filter = {
                    'report__year': year, 
                    'report__is_published': True,
                    'status__in': ['A', 'D'] if show_ter_status else ['D'],  # Учитываем отчеты на согласовании
                    'school__closter': cluster
                }
                
                # Если указаны ТУ/ДО, добавляем их в фильтр
                if ter_admin_ids:
                    school_reports_filter['school__ter_admin_id__in'] = ter_admin_ids
                
                school_reports = SchoolReport.objects.filter(**school_reports_filter).select_related('school')
                
                # Подсчитываем количество отчетов в каждой зоне для текущего кластера и года
                red_zone = 0
                yellow_zone = 0
                green_zone = 0
                total = 0
                
                # Для каждого отчета школы определяем общую зону отчета
                for sr in school_reports:
                    # Пропускаем отчеты без школы или с архивной школой
                    if not sr.school or sr.school.is_archived:
                        continue
                    
                    if sr.zone == 'R':
                        red_zone += 1
                    elif sr.zone == 'Y':
                        yellow_zone += 1
                    elif sr.zone == 'G':
                        green_zone += 1
                    total += 1
                
                if total > 0:
                    cluster_years_stats.append({
                        'year': year.year,
                        'red_zone': red_zone,
                        'yellow_zone': yellow_zone,
                        'green_zone': green_zone,
                        'total': total,
                        'red_percent': f"{(red_zone / total) * 100:.1f}".replace(',', '.'),
                        'yellow_percent': f"{(yellow_zone / total) * 100:.1f}".replace(',', '.'),
                        'green_percent': f"{(green_zone / total) * 100:.1f}".replace(',', '.')
                    })
            
            if cluster_years_stats:
                # Группируем данные по кластерам
                cluster_stats.append({
                    'cluster_name': cluster.name,
                    'years': cluster_years_stats,
                    'rowspan': len(cluster_years_stats)
                })
    
    return cluster_stats

@login_required
@csrf_exempt
def get_schools_by_ter_admin(request):
    """
    AJAX endpoint для получения школ по выбранному ТУ/ДО
    """
    if request.method == 'GET':
        ter_admin_id = request.GET.get('ter_admin_id')
        
        if ter_admin_id:
            try:
                schools = School.objects.filter(
                    ter_admin_id=ter_admin_id, 
                    is_archived=False
                ).order_by('name')
                
                schools_data = []
                for school in schools:
                    schools_data.append({
                        'id': school.id,
                        'name': str(school)
                    })
                
                return JsonResponse({
                    'status': 'success',
                    'schools': schools_data
                })
            except Exception as e:
                return JsonResponse({
                    'status': 'error',
                    'message': str(e)
                })
        else:
            # Если ТУ/ДО не выбрано, возвращаем все школы
            ter_admins = TerAdmin.objects.filter(representatives=request.user)
            if not ter_admins.exists():
                ter_admins = TerAdmin.objects.all()
            
            schools_data = []
            for ter_admin in ter_admins:
                for school in ter_admin.schools.filter(is_archived=False):
                    schools_data.append({
                        'id': school.id,
                        'name': str(school)
                    })
            
            return JsonResponse({
                'status': 'success',
                'schools': schools_data
            })
    
    return JsonResponse({
        'status': 'error',
        'message': 'Invalid request method'
    })

# Диагностическая функция для отладки БЕЗ декоратора авторизации
def debug_auth_status(request):
    """Диагностическая функция для проверки авторизации БЕЗ @login_required"""
    debug_info = []
    
    try:
        debug_info.append("=== ДИАГНОСТИКА АВТОРИЗАЦИИ (БЕЗ @login_required) ===")
        debug_info.append(f"request.user: {request.user}")
        debug_info.append(f"request.user type: {type(request.user)}")
        
        if hasattr(request.user, 'is_authenticated'):
            debug_info.append(f"request.user.is_authenticated: {request.user.is_authenticated}")
            if callable(request.user.is_authenticated):
                debug_info.append(f"request.user.is_authenticated(): {request.user.is_authenticated()}")
        else:
            debug_info.append("request.user.is_authenticated: ОТСУТСТВУЕТ")
            
        debug_info.append(f"request.user.is_anonymous: {getattr(request.user, 'is_anonymous', 'N/A')}")
        debug_info.append(f"request.user.id: {getattr(request.user, 'id', 'N/A')}")
        debug_info.append(f"request.user.username: {getattr(request.user, 'username', 'N/A')}")
        debug_info.append(f"request.user.email: {getattr(request.user, 'email', 'N/A')}")
        debug_info.append(f"request.user.is_superuser: {getattr(request.user, 'is_superuser', 'N/A')}")
        debug_info.append(f"request.user.is_staff: {getattr(request.user, 'is_staff', 'N/A')}")
        debug_info.append(f"request.user.is_active: {getattr(request.user, 'is_active', 'N/A')}")
        
        # Проверка сессии
        debug_info.append(f"Session exists: {hasattr(request, 'session')}")
        if hasattr(request, 'session'):
            debug_info.append(f"Session key: {request.session.session_key}")
            debug_info.append(f"Session is_empty: {request.session.is_empty()}")
            debug_info.append(f"Session data keys: {list(request.session.keys())}")
            debug_info.append(f"USER_ID in session: {request.session.get('_auth_user_id', 'NOT FOUND')}")
            debug_info.append(f"USER_BACKEND in session: {request.session.get('_auth_user_backend', 'NOT FOUND')}")
            debug_info.append(f"USER_HASH in session: {request.session.get('_auth_user_hash', 'NOT FOUND')}")
        
        # Проверяем настройки Django
        from django.conf import settings
        debug_info.append("=== НАСТРОЙКИ DJANGO ===")
        debug_info.append(f"AUTH_USER_MODEL: {getattr(settings, 'AUTH_USER_MODEL', 'DEFAULT')}")
        debug_info.append(f"SESSION_ENGINE: {getattr(settings, 'SESSION_ENGINE', 'NOT SET')}")
        debug_info.append(f"SESSION_COOKIE_NAME: {getattr(settings, 'SESSION_COOKIE_NAME', 'NOT SET')}")
        
        # Проверяем middleware
        middleware_list = getattr(settings, 'MIDDLEWARE', [])
        debug_info.append(f"MIDDLEWARE содержит AuthenticationMiddleware: {'django.contrib.auth.middleware.AuthenticationMiddleware' in middleware_list}")
        debug_info.append(f"MIDDLEWARE содержит SessionMiddleware: {'django.contrib.sessions.middleware.SessionMiddleware' in middleware_list}")
        
        # Диагностика Redis кеша
        debug_info.append("=== ДИАГНОСТИКА REDIS КЕША ===")
        try:
            from django.core.cache import cache
            debug_info.append("Проверка подключения к Redis...")
            
            # Тестовый ключ
            test_key = "django_test_connection"
            test_value = "test_value_123"
            
            # Попытка записи и чтения
            cache.set(test_key, test_value, timeout=10)
            retrieved_value = cache.get(test_key)
            
            if retrieved_value == test_value:
                debug_info.append("✅ Redis кеш работает корректно")
                cache.delete(test_key)  # Удаляем тестовый ключ
            else:
                debug_info.append(f"❌ Redis кеш НЕ работает: записано '{test_value}', получено '{retrieved_value}'")
                
            # Дополнительная информация о кеше
            cache_info = cache._cache.get_client().info()
            debug_info.append(f"Redis версия: {cache_info.get('redis_version', 'N/A')}")
            debug_info.append(f"Redis подключения: {cache_info.get('connected_clients', 'N/A')}")
            debug_info.append(f"Redis используемая память: {cache_info.get('used_memory_human', 'N/A')}")
            
        except Exception as e:
            debug_info.append(f"❌ Ошибка подключения к Redis: {e}")
        
        # Попытка получить пользователя из базы
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            debug_info.append(f"Кастомная модель пользователя: {User}")
            
            # Проверяем, есть ли суперпользователи
            superusers = User.objects.filter(is_superuser=True)
            debug_info.append(f"Суперпользователи в системе: {list(superusers.values_list('username', 'email', 'is_active'))}")
            
        except Exception as e:
            debug_info.append(f"Ошибка при работе с моделью пользователя: {e}")
            
        # Проверяем cookies
        debug_info.append("=== COOKIES ===")
        session_cookie_name = getattr(settings, 'SESSION_COOKIE_NAME', 'sessionid')
        debug_info.append(f"Session cookie ({session_cookie_name}): {request.COOKIES.get(session_cookie_name, 'NOT FOUND')}")
        debug_info.append(f"CSRF cookie: {request.COOKIES.get('csrftoken', 'NOT FOUND')}")
        debug_info.append(f"Все cookies: {list(request.COOKIES.keys())}")
        
        # Проверяем переменные окружения
        debug_info.append("=== ПЕРЕМЕННЫЕ ОКРУЖЕНИЯ ===")
        import os
        debug_info.append(f"DEBUG env: {os.environ.get('DEBUG', 'NOT SET')}")
        debug_info.append(f"SECRET_KEY env (первые 10 символов): {(os.environ.get('SECRET_KEY', 'NOT SET'))[:10]}...")
        debug_info.append(f"DATABASE_NAME env: {os.environ.get('DATABASE_NAME', 'NOT SET')}")
        debug_info.append(f"DATABASE_HOST env: {os.environ.get('DATABASE_HOST', 'NOT SET')}")
        
    except Exception as e:
        debug_info.append(f"Ошибка при диагностике: {str(e)}")
        import traceback
        debug_info.append(f"Traceback: {traceback.format_exc()}")
        
    return render(request, "dashboards/debug.html", {"debug_info": debug_info})


# Диагностическая функция для отладки
